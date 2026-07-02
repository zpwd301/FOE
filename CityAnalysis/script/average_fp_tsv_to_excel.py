#!/usr/bin/env python3
"""Convert an average-expected-FP TSV report into a styled Excel workbook."""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NUMERIC_COLUMNS = {"Average Expected FP", "Base FP", "Level", "Count", "Passive FP Boost"}
REPORT_SUFFIX = "_Blue_Galaxy_Collection_Recommendation_Report"
REPORT_EXPLANATION = (
    "This report ranks city buildings for Blue Galaxy collection priority by estimated Forge Point value; "
    "Average Expected FP assumes the building is motivated and applies the city's passive FP boost, while Base FP "
    "shows the unboosted expected FP from the best listed collection option."
)
HEADER_ROW = 2
DATA_START_ROW = 3
ROW_BAND_FILLS = ("F7F3E8", "E8F1EA")
MALLARD_GREEN = "214E34"
MALLARD_BLUE = "516C8B"
MALLARD_CHESTNUT = "7D4B35"
MALLARD_BILL = "D9A441"
MALLARD_CREAM = "F7F3E8"
GRID_COLOR = "B7C7BA"


def default_input_path(output_dir: Path) -> Path:
    candidates = sorted(
        output_dir.rglob(f"*{REPORT_SUFFIX}.tsv"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]
    return output_dir / f"Sel{REPORT_SUFFIX}.tsv"


def parse_args() -> argparse.Namespace:
    base_dir = Path(__file__).resolve().parents[1]
    output_dir = base_dir / "output"
    parser = argparse.ArgumentParser(
        description="Convert a sorted average-expected-FP TSV report into a styled Excel workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_input_path(output_dir),
        help="Path to the TSV report. Defaults to the newest matching output file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path. Defaults to the TSV path with an .xlsx suffix.",
    )
    parser.add_argument(
        "--plain",
        action="store_true",
        help="Use simple readable formatting without the report color scheme or gradient.",
    )
    return parser.parse_args()


def load_tsv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        return list(reader)


def excel_value(column_name: str, raw_value: str):
    if column_name not in NUMERIC_COLUMNS:
        return raw_value
    if raw_value == "":
        return ""
    if "." in raw_value:
        return float(raw_value)
    return int(raw_value)


def sheet_title_from_path(path: Path) -> str:
    stem = path.stem
    if stem.endswith(REPORT_SUFFIX):
        stem = stem[: -len(REPORT_SUFFIX)]
    cleaned = stem.replace("_", " ").replace("-", " ").strip() or "FP Report"
    return cleaned[:31]


def apply_styles(sheet, headers: List[str], row_count: int) -> None:
    note_font = Font(name="Calibri", size=11, italic=True, color="1F1F1F")
    note_fill = PatternFill(fill_type="solid", start_color=MALLARD_CREAM, end_color=MALLARD_CREAM)
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color=MALLARD_GREEN, end_color=MALLARD_GREEN)
    body_font = Font(name="Calibri", size=11, color="1F1F1F")
    name_font = Font(name="Calibri", size=11, bold=True, color=MALLARD_GREEN)
    production_font = Font(name="Calibri", size=11, color=MALLARD_CHESTNUT)
    light_value_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color=GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    sheet.freeze_panes = f"A{DATA_START_ROW}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{row_count + HEADER_ROW}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = MALLARD_BLUE

    note_cell = sheet.cell(row=1, column=1)
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.border = border
    note_cell.alignment = left
    sheet.row_dimensions[1].height = 34

    for cell in sheet[HEADER_ROW]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    header_index = {name: idx + 1 for idx, name in enumerate(headers)}
    production_col = header_index.get("Production")
    name_col = header_index.get("Name")
    avg_col = header_index.get("Average Expected FP")

    for row_idx in range(DATA_START_ROW, row_count + DATA_START_ROW):
        band_fill = PatternFill(
            fill_type="solid",
            start_color=ROW_BAND_FILLS[(row_idx - DATA_START_ROW) % 2],
            end_color=ROW_BAND_FILLS[(row_idx - DATA_START_ROW) % 2],
        )
        sheet.row_dimensions[row_idx].height = 30
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.fill = band_fill
            cell.border = border
            cell.alignment = right if headers[col_idx - 1] in NUMERIC_COLUMNS else center

        if name_col is not None:
            name_cell = sheet.cell(row=row_idx, column=name_col)
            name_cell.font = name_font
            name_cell.alignment = left

        if production_col is not None:
            production_cell = sheet.cell(row=row_idx, column=production_col)
            production_cell.font = production_font
            production_cell.alignment = left

    if avg_col is not None and row_count > 0:
        start = f"{get_column_letter(avg_col)}{DATA_START_ROW}"
        end = f"{get_column_letter(avg_col)}{row_count + HEADER_ROW}"
        avg_range = f"{start}:{end}"
        sheet.conditional_formatting.add(
            avg_range,
            ColorScaleRule(
                start_type="min",
                start_color=MALLARD_CREAM,
                mid_type="percentile",
                mid_value=60,
                mid_color=MALLARD_BILL,
                end_type="max",
                end_color=MALLARD_BLUE,
            ),
        )
        # High-value cells land on the darkest end of the scale, so switch them to white text.
        sheet.conditional_formatting.add(
            avg_range,
            Rule(
                type="expression",
                formula=[
                    f"{start}>=(PERCENTILE(${get_column_letter(avg_col)}${DATA_START_ROW}:${get_column_letter(avg_col)}${row_count + HEADER_ROW},0.85))"
                ],
                dxf=DifferentialStyle(font=light_value_font),
            ),
        )

    top_fill = PatternFill(fill_type="solid", start_color="E1EBDD", end_color="E1EBDD")
    for row_idx in range(DATA_START_ROW, min(row_count + DATA_START_ROW, DATA_START_ROW + 5)):
        for col_idx in range(1, len(headers) + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = top_fill


def apply_plain_styles(sheet, headers: List[str], row_count: int) -> None:
    note_font = Font(name="Calibri", size=11, italic=True, color="1F1F1F")
    note_fill = PatternFill(fill_type="solid", start_color="F7F7F7", end_color="F7F7F7")
    header_font = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
    header_fill = PatternFill(fill_type="solid", start_color="EDEDED", end_color="EDEDED")
    body_font = Font(name="Calibri", size=11, color="1F1F1F")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top", wrap_text=True)

    sheet.freeze_panes = f"A{DATA_START_ROW}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{row_count + HEADER_ROW}"

    note_cell = sheet.cell(row=1, column=1)
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.border = border
    note_cell.alignment = left
    sheet.row_dimensions[1].height = 34

    for cell in sheet[HEADER_ROW]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row_idx in range(DATA_START_ROW, row_count + DATA_START_ROW):
        sheet.row_dimensions[row_idx].height = 45
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.border = border
            if header in NUMERIC_COLUMNS:
                cell.alignment = right
            elif header == "Production":
                cell.alignment = left
            else:
                cell.alignment = center


def adjust_widths(sheet, headers: List[str], rows: List[Dict[str, str]]) -> None:
    width_limits = {
        "Average Expected FP": 20,
        "Base FP": 12,
        "Name": 34,
        "Level": 9,
        "Age": 9,
        "Count": 9,
        "Passive FP Boost": 17,
        "Production": 100,
    }
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(header)
        for row in rows:
            max_length = max(max_length, len(str(row.get(header, ""))))
        padding = 2
        width = min(max_length + padding, width_limits.get(header, 40))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(rows: List[Dict[str, str]], output_path: Path, source_path: Path, *, plain: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = list(rows[0].keys()) if rows else [
        "Average Expected FP",
        "Base FP",
        "Name",
        "Level",
        "Age",
        "Count",
        "Passive FP Boost",
        "Production",
    ]
    sheet.title = sheet_title_from_path(source_path)
    sheet.append([REPORT_EXPLANATION])
    if headers:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.append(headers)

    for row in rows:
        sheet.append([excel_value(header, row.get(header, "")) for header in headers])

    if plain:
        apply_plain_styles(sheet, headers, len(rows))
    else:
        apply_styles(sheet, headers, len(rows))
    adjust_widths(sheet, headers, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = (
        args.output.expanduser().resolve()
        if args.output is not None
        else input_path.with_suffix(".xlsx")
    )
    rows = load_tsv(input_path)
    write_workbook(rows, output_path, input_path, plain=args.plain)


if __name__ == "__main__":
    main()
