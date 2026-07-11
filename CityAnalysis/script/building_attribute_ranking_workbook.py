#!/usr/bin/env python3
"""Build a configurable building ranking workbook from reference CityEntities."""
from __future__ import annotations

import argparse
import json
import math
import os
import tempfile
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from building_ranking_model import *  # noqa: F401,F403
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


def apply_building_name_color_rules(
    sheet: Any,
    data_start: int,
    data_end: int,
    entity_col: int,
    event_abbreviations_: Sequence[str],
) -> None:
    if data_end < data_start:
        return

    entity_letter = get_column_letter(entity_col)
    entity_cell = f"${entity_letter}{data_start}"
    name_range = f"A{data_start}:A{data_end}"

    rules = [
        (f'LEFT({entity_cell},{len(GBG_REWARD_PREFIX)})="{GBG_REWARD_PREFIX}"', REWARD_GROUP_COLORS["GBG"]),
        (f'LEFT({entity_cell},{len(QI_REWARD_PREFIX)})="{QI_REWARD_PREFIX}"', REWARD_GROUP_COLORS["QI"]),
        (
            "OR("
            + ",".join(
                f'LEFT({entity_cell},{len(prefix)})="{prefix}"'
                for prefix in GE_REWARD_PREFIXES
            )
            + ")",
            REWARD_GROUP_COLORS["GE"],
        ),
    ]

    excluded_prefix_checks = [
        f'LEFT({entity_cell},{len(GBG_REWARD_PREFIX)})<>"{GBG_REWARD_PREFIX}"',
        f'LEFT({entity_cell},{len(QI_REWARD_PREFIX)})<>"{QI_REWARD_PREFIX}"',
        *[
            f'LEFT({entity_cell},{len(prefix)})<>"{prefix}"'
            for prefix in GE_REWARD_PREFIXES
        ],
    ]
    for abbreviation, color in event_color_map(event_abbreviations_).items():
        prefix = f"{MULTI_AGE_PREFIX}{abbreviation}{current_year_suffix()}"
        formula = (
            "AND("
            + ",".join(
                [
                    *excluded_prefix_checks,
                    f'LEFT({entity_cell},{len(prefix)})="{prefix}"',
                ]
            )
            + ")"
        )
        rules.append((formula, color))

    for formula, color in rules:
        sheet.conditional_formatting.add(
            name_range,
            FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
        )


def overall_weight_cell_value(row_idx: int) -> str:
    raw_col = get_column_letter(OVERALL_RAW_WEIGHT_COLUMN)
    group_col = get_column_letter(OVERALL_WEIGHT_GROUP_COLUMN)
    budget_col = get_column_letter(OVERALL_WEIGHT_BUDGET_COLUMN)
    raw_cell = f"{raw_col}{row_idx}"
    group_cell = f"{group_col}{row_idx}"
    budget_cell = f"{budget_col}{row_idx}"
    return (
        f'=IF({raw_cell}=0,0,{raw_cell}*IFERROR({budget_cell}/'
        f'SUMIF(${group_col}:${group_col},{group_cell},${raw_col}:${raw_col}),0))'
    )


def formula_expr(value: Any) -> str:
    if isinstance(value, str) and value.startswith("="):
        return value[1:]
    numeric = as_float(value)
    return cached_number(numeric if numeric is not None else 0.0)


def weight_mode_formula(default_value: Any, override_col: int, row_idx: int) -> str:
    default_expr = formula_expr(default_value)
    override_cell = f"{get_column_letter(override_col)}{row_idx}"
    return (
        f'=IF({ADVANCED_WEIGHT_MODE_CELL}="Default",{default_expr},'
        f'IF(ISBLANK({override_cell}),{default_expr},{override_cell}))'
    )


def age_data_lookup_formula(
    entity_id: str,
    column_letter: str,
    max_row: int,
    default_value: str = "0",
) -> str:
    lookup_key = f"{excel_string(entity_id + '|')}&{CONTROLS_SHEET_REF}!{CITY_AGE_CELL}"
    return (
        f"IFERROR(INDEX({AGE_DATA_SHEET_REF}!${column_letter}$2:${column_letter}${max_row},"
        f"MATCH({lookup_key},{AGE_DATA_SHEET_REF}!$A$2:$A${max_row},0)),{default_value})"
    )


def cached_number(value: float) -> str:
    if math.isclose(value, 0.0, abs_tol=1e-12):
        return "0"
    return f"{value:.12g}"


def set_formula_cache(root: ET.Element, cell_ref: str, value: float) -> None:
    namespace = {"main": XLSX_MAIN_NS}
    cell = root.find(f".//main:c[@r='{cell_ref}']", namespace)
    if cell is None:
        return
    for value_node in list(cell.findall("main:v", namespace)):
        cell.remove(value_node)
    value_node = ET.SubElement(cell, f"{{{XLSX_MAIN_NS}}}v")
    value_node.text = cached_number(value)


def worksheet_filenames_by_title(workbook_zip: ZipFile) -> Dict[str, str]:
    namespace = {"main": XLSX_MAIN_NS, "rel": PACKAGE_REL_NS}
    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("rel:Relationship", namespace)
        if "Id" in rel.attrib and "Target" in rel.attrib
    }

    filenames: Dict[str, str] = {}
    for sheet in workbook_root.findall("main:sheets/main:sheet", namespace):
        title = sheet.attrib.get("name")
        relation_id = sheet.attrib.get(f"{{{XLSX_REL_NS}}}id")
        target = rel_targets.get(relation_id or "")
        if not title or not target:
            continue
        if target.startswith("/"):
            filename = target.lstrip("/")
        elif target.startswith("xl/"):
            filename = target
        else:
            filename = f"xl/{target}"
        filenames[title] = filename
    return filenames


def populate_formula_caches(
    output_file: str,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
    stats: Dict[str, Dict[str, float]],
) -> None:
    if not records:
        return

    ET.register_namespace("", XLSX_MAIN_NS)
    overall_weights = overall_weight_map(attr_keys)
    overall_weight_func = lambda key: overall_weights.get(key, 0.0)
    coefficients, offsets, total_weight = scoring_terms(attr_keys, stats, overall_weight_func)
    fighting_coefficients, fighting_offsets, fighting_total_weight = scoring_terms(attr_keys, stats, fighting_weight_for_attr)
    fp_goods_coefficients, fp_goods_offsets, fp_goods_total_weight = scoring_terms(attr_keys, stats, fp_goods_weight_for_attr)
    qi_coefficients, qi_offsets, qi_total_weight = scoring_terms(attr_keys, stats, qi_weight_for_attr)
    scores = [formula_score_from_terms(record, attr_keys, coefficients, offsets, total_weight) for record in records]
    fighting_scores = [
        formula_score_from_terms(record, attr_keys, fighting_coefficients, fighting_offsets, fighting_total_weight)
        for record in records
    ]
    fp_goods_scores = [
        formula_score_from_terms(record, attr_keys, fp_goods_coefficients, fp_goods_offsets, fp_goods_total_weight)
        for record in records
    ]
    qi_scores = [
        formula_score_from_terms(record, attr_keys, qi_coefficients, qi_offsets, qi_total_weight)
        for record in records
    ]
    adjusted_areas = [adjusted_area(record) for record in records]
    overall_efficiency_scores = [
        score / area if not math.isclose(area, 0.0) else 0.0
        for score, area in zip(scores, adjusted_areas)
    ]
    fighting_efficiency_scores = [
        score / area if not math.isclose(area, 0.0) else 0.0
        for score, area in zip(fighting_scores, adjusted_areas)
    ]
    fp_goods_efficiency_scores = [
        score / area if not math.isclose(area, 0.0) else 0.0
        for score, area in zip(fp_goods_scores, adjusted_areas)
    ]
    qi_efficiency_scores = [
        score / area if not math.isclose(area, 0.0) else 0.0
        for score, area in zip(qi_scores, adjusted_areas)
    ]
    ranks = [1 + sum(1 for other_score in scores if other_score > score) for score in scores]
    fighting_ranks = [
        1 + sum(1 for other_score in fighting_scores if other_score > score)
        for score in fighting_scores
    ]
    fp_goods_ranks = [
        1 + sum(1 for other_score in fp_goods_scores if other_score > score)
        for score in fp_goods_scores
    ]
    qi_ranks = [
        1 + sum(1 for other_score in qi_scores if other_score > score)
        for score in qi_scores
    ]
    overall_efficiency_ranks = [
        1 + sum(1 for other_score in overall_efficiency_scores if other_score > score)
        for score in overall_efficiency_scores
    ]
    fighting_efficiency_ranks = [
        1 + sum(1 for other_score in fighting_efficiency_scores if other_score > score)
        for score in fighting_efficiency_scores
    ]
    fp_goods_efficiency_ranks = [
        1 + sum(1 for other_score in fp_goods_efficiency_scores if other_score > score)
        for score in fp_goods_efficiency_scores
    ]
    qi_efficiency_ranks = [
        1 + sum(1 for other_score in qi_efficiency_scores if other_score > score)
        for score in qi_efficiency_scores
    ]
    attr_columns = {key: RAW_START_COLUMN + idx for idx, key in enumerate(attr_keys)}

    tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(output_file))
    os.close(tmp_fd)
    try:
        with ZipFile(output_file, "r") as source, ZipFile(tmp_name, "w", ZIP_DEFLATED) as target:
            sheet_files = worksheet_filenames_by_title(source)
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == sheet_files.get(ADVANCED_CONTROLS_SHEET):
                    root = ET.fromstring(data)
                    set_formula_cache(root, OVERALL_TOTAL_WEIGHT_CELL.replace("$", ""), total_weight)
                    set_formula_cache(root, FIGHTING_TOTAL_WEIGHT_CELL.replace("$", ""), fighting_total_weight)
                    set_formula_cache(root, FP_GOODS_TOTAL_WEIGHT_CELL.replace("$", ""), fp_goods_total_weight)
                    set_formula_cache(root, QI_TOTAL_WEIGHT_CELL.replace("$", ""), qi_total_weight)
                    for idx, key in enumerate(attr_keys):
                        row = WEIGHT_START_ROW + idx
                        set_formula_cache(root, f"C{row}", overall_weights.get(key, 0.0))
                        set_formula_cache(root, f"D{row}", abs(overall_weights.get(key, 0.0)))
                        if isinstance(overall_raw_weight_cell_value(key), str):
                            set_formula_cache(root, f"{get_column_letter(OVERALL_RAW_WEIGHT_COLUMN)}{row}", overall_raw_weight_for_attr(key))
                        set_formula_cache(root, f"E{row}", fighting_weight_for_attr(key))
                        set_formula_cache(root, f"F{row}", abs(fighting_weight_for_attr(key)))
                        set_formula_cache(root, f"G{row}", fp_goods_weight_for_attr(key))
                        set_formula_cache(root, f"H{row}", abs(fp_goods_weight_for_attr(key)))
                        set_formula_cache(root, f"I{row}", qi_weight_for_attr(key))
                        set_formula_cache(root, f"J{row}", abs(qi_weight_for_attr(key)))
                        if key in {PROD_FP_ATTR, PROD_GOODS_ATTR, PROD_GUILD_GOODS_ATTR, PROD_MEDALS_ATTR}:
                            set_formula_cache(root, f"L{row}", stats[key]["min"])
                            set_formula_cache(root, f"M{row}", stats[key]["max"])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(OVERALL_SOURCE_SHEET):
                    root = ET.fromstring(data)
                    raw_start = RAW_START_COLUMN
                    for idx, value in enumerate(coefficients):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}2", value)
                    for idx, value in enumerate(offsets):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}3", value)
                    for idx, value in enumerate(fighting_coefficients):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}4", value)
                    for idx, value in enumerate(fighting_offsets):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}5", value)
                    for idx, value in enumerate(fp_goods_coefficients):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}6", value)
                    for idx, value in enumerate(fp_goods_offsets):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}7", value)
                    for idx, value in enumerate(qi_coefficients):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}8", value)
                    for idx, value in enumerate(qi_offsets):
                        set_formula_cache(root, f"{get_column_letter(raw_start + idx)}9", value)
                    for idx, score in enumerate(scores):
                        row = BUILDING_DATA_START_ROW + idx
                        set_formula_cache(root, f"B{row}", ranks[idx])
                        set_formula_cache(root, f"C{row}", score)
                        for key in (PROD_FP_ATTR, PROD_GOODS_ATTR, PROD_GUILD_GOODS_ATTR, PROD_MEDALS_ATTR):
                            if key in attr_columns:
                                set_formula_cache(
                                    root,
                                    f"{get_column_letter(attr_columns[key])}{row}",
                                    effective_attr_value(records[idx], key),
                                )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(OVERALL_SCORE_SHEET):
                    root = ET.fromstring(data)
                    for idx, score in enumerate(scores, start=2):
                        set_formula_cache(root, f"A{idx}", BUILDING_DATA_START_ROW + idx - 2)
                        set_formula_cache(root, f"B{idx}", score)
                        set_formula_cache(root, f"C{idx}", ranks[idx - 2])
                        set_formula_cache(root, f"D{idx}", adjusted_areas[idx - 2])
                        set_formula_cache(root, f"E{idx}", overall_efficiency_scores[idx - 2])
                        set_formula_cache(root, f"F{idx}", overall_efficiency_ranks[idx - 2])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(OVERALL_EFFICIENCY_SHEET):
                    root = ET.fromstring(data)
                    display_attr_keys = [
                        key
                        for key in overall_ranking_display_attr_keys(attr_keys)
                        if not is_road_connection_attr_key(key)
                    ]
                    attr_start = 9
                    source_row_col = attr_start + len(display_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-overall_efficiency_scores[record_idx], record_idx),
                    )[:OVERALL_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", overall_efficiency_scores[record_idx])
                        set_formula_cache(root, f"D{output_idx}", scores[record_idx])
                        set_formula_cache(root, f"G{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"H{output_idx}", adjusted_areas[record_idx])
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(display_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                display_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(FIGHTING_SCORE_SHEET):
                    root = ET.fromstring(data)
                    for idx, score in enumerate(fighting_scores, start=2):
                        set_formula_cache(root, f"A{idx}", BUILDING_DATA_START_ROW + idx - 2)
                        set_formula_cache(root, f"B{idx}", score)
                        set_formula_cache(root, f"C{idx}", fighting_ranks[idx - 2])
                        set_formula_cache(root, f"D{idx}", adjusted_areas[idx - 2])
                        set_formula_cache(root, f"E{idx}", fighting_efficiency_scores[idx - 2])
                        set_formula_cache(root, f"F{idx}", fighting_efficiency_ranks[idx - 2])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get("Fighting Ranking"):
                    root = ET.fromstring(data)
                    fighting_attr_keys = [key for key in attr_keys if fighting_weight_for_attr(key)]
                    attr_start = 7
                    source_row_col = attr_start + len(fighting_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-fighting_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", fighting_scores[record_idx])
                        set_formula_cache(root, f"F{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(fighting_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(FIGHTING_EFFICIENCY_SHEET):
                    root = ET.fromstring(data)
                    fighting_attr_keys = [key for key in attr_keys if fighting_weight_for_attr(key)]
                    attr_start = 9
                    source_row_col = attr_start + len(fighting_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-fighting_efficiency_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", fighting_efficiency_scores[record_idx])
                        set_formula_cache(root, f"D{output_idx}", fighting_scores[record_idx])
                        set_formula_cache(root, f"G{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"H{output_idx}", adjusted_areas[record_idx])
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(fighting_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(FP_GOODS_SCORE_SHEET):
                    root = ET.fromstring(data)
                    for idx, score in enumerate(fp_goods_scores, start=2):
                        set_formula_cache(root, f"A{idx}", BUILDING_DATA_START_ROW + idx - 2)
                        set_formula_cache(root, f"B{idx}", score)
                        set_formula_cache(root, f"C{idx}", fp_goods_ranks[idx - 2])
                        set_formula_cache(root, f"D{idx}", adjusted_areas[idx - 2])
                        set_formula_cache(root, f"E{idx}", fp_goods_efficiency_scores[idx - 2])
                        set_formula_cache(root, f"F{idx}", fp_goods_efficiency_ranks[idx - 2])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(FP_GOODS_PRODUCTION_SHEET):
                    root = ET.fromstring(data)
                    production_attr_keys = fp_goods_display_attr_keys(attr_keys)
                    attr_start = 7
                    source_row_col = attr_start + len(production_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-fp_goods_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", fp_goods_scores[record_idx])
                        set_formula_cache(root, f"F{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(production_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(FP_GOODS_EFFICIENCY_SHEET):
                    root = ET.fromstring(data)
                    production_attr_keys = fp_goods_display_attr_keys(attr_keys)
                    attr_start = 9
                    source_row_col = attr_start + len(production_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-fp_goods_efficiency_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", fp_goods_efficiency_scores[record_idx])
                        set_formula_cache(root, f"D{output_idx}", fp_goods_scores[record_idx])
                        set_formula_cache(root, f"G{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"H{output_idx}", adjusted_areas[record_idx])
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(production_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(QI_SCORE_SHEET):
                    root = ET.fromstring(data)
                    for idx, score in enumerate(qi_scores, start=2):
                        set_formula_cache(root, f"A{idx}", BUILDING_DATA_START_ROW + idx - 2)
                        set_formula_cache(root, f"B{idx}", score)
                        set_formula_cache(root, f"C{idx}", qi_ranks[idx - 2])
                        set_formula_cache(root, f"D{idx}", adjusted_areas[idx - 2])
                        set_formula_cache(root, f"E{idx}", qi_efficiency_scores[idx - 2])
                        set_formula_cache(root, f"F{idx}", qi_efficiency_ranks[idx - 2])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(QI_RANKING_SHEET):
                    root = ET.fromstring(data)
                    qi_attr_keys = qi_display_attr_keys(attr_keys)
                    attr_start = 7
                    source_row_col = attr_start + len(qi_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-qi_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", qi_scores[record_idx])
                        set_formula_cache(root, f"F{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(qi_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                elif item.filename == sheet_files.get(QI_EFFICIENCY_SHEET):
                    root = ET.fromstring(data)
                    qi_attr_keys = qi_display_attr_keys(attr_keys)
                    attr_start = 9
                    source_row_col = attr_start + len(qi_attr_keys) + 4
                    top_indices = sorted(
                        range(len(records)),
                        key=lambda record_idx: (-qi_efficiency_scores[record_idx], record_idx),
                    )[:FIGHTING_TOP_N]
                    for output_idx, record_idx in enumerate(top_indices, start=5):
                        source_row = BUILDING_DATA_START_ROW + record_idx
                        record = records[record_idx]
                        set_formula_cache(root, f"B{output_idx}", output_idx - 4)
                        set_formula_cache(root, f"C{output_idx}", qi_efficiency_scores[record_idx])
                        set_formula_cache(root, f"D{output_idx}", qi_scores[record_idx])
                        set_formula_cache(root, f"G{output_idx}", float(record["area"] or 0))
                        set_formula_cache(root, f"H{output_idx}", adjusted_areas[record_idx])
                        set_formula_cache(root, f"{get_column_letter(source_row_col)}{output_idx}", source_row)
                        for attr_idx, key in enumerate(qi_attr_keys, start=attr_start):
                            set_formula_cache(
                                root,
                                f"{get_column_letter(attr_idx)}{output_idx}",
                                effective_attr_value(record, key),
                            )
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                target.writestr(item, data)
        os.replace(tmp_name, output_file)
        os.chmod(output_file, 0o644)
    except Exception:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)
        raise


def write_controls_sheet(
    workbook: Workbook,
    reference_file: str,
    era: str,
    available_only: bool,
    all_ages: bool = False,
    category_options: Sequence[str] = (),
) -> None:
    sheet = workbook.active
    sheet.title = CONTROLS_SHEET
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 115

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    editable_fill = PatternFill("solid", fgColor=EDITABLE_FILL_COLOR)
    slider_fill = PatternFill("solid", fgColor=SLIDER_FILL_COLOR)
    slider_selected_fill = PatternFill("solid", fgColor=SLIDER_SELECTED_FILL_COLOR)
    context_fill = PatternFill("solid", fgColor=CONTROL_CONTEXT_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    track_side = Side(style="medium", color="79A878")
    no_side = Side(style=None)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"] = (
        "Interactive Building Rankings - All Ages - Main Controls"
        if all_ages
        else "Building Attribute Ranking Main Controls - Scale Values 1-5"
    )
    sheet["A1"].font = Font(bold=True, size=18, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A1:G1")
    sheet.row_dimensions[1].height = 30

    rows = [
        (
            2,
            "How to use",
            "Start here: pick your city age, select a building source category filter, then enter your city’s estimated base production for FP, regular goods, guild goods, medals, and special goods in the yellow cells. These values should not include any percentage-based boosts.\n\nNext, choose the Fighting, QI, and FP/Goods Production focus settings that best match your priorities. For each 1–5 scale, 1 favors the left label, 3 is balanced, and 5 favors the right label.\n\nThen review the ranking sheets.",
            PatternFill(fill_type=None),
        ),
        (3, "Select Your City Age" if all_ages else "Assumed age", selected_age_display(era, all_ages), editable_fill if all_ages else context_fill),
        (5, "Building source category filter", ALL_BUILDING_CATEGORIES, editable_fill),
        (6, "Estimated base FP production", DEFAULT_ESTIMATED_FP_PRODUCTION, editable_fill),
        (7, "Estimated base regular goods production", DEFAULT_ESTIMATED_GOODS_PRODUCTION, editable_fill),
        (8, "Estimated base guild goods production", DEFAULT_ESTIMATED_GUILD_GOODS_PRODUCTION, editable_fill),
        (9, "Estimated base medal production", DEFAULT_ESTIMATED_MEDAL_PRODUCTION, editable_fill),
        (10, "QI fighter role", "Both", editable_fill),
        (11, "Estimated base special goods production", DEFAULT_ESTIMATED_SPECIAL_GOODS_PRODUCTION, editable_fill),
    ]
    for row_idx, label, value, fill in rows:
        sheet.cell(row_idx, 1, label)
        sheet.cell(row_idx, 2, value)
        sheet.cell(row_idx, 1).font = Font(size=12, color=HEADER_FONT_COLOR)
        sheet.cell(row_idx, 2).font = Font(size=12, color=TITLE_FONT_COLOR)
        sheet.cell(row_idx, 1).border = border
        sheet.cell(row_idx, 2).border = border
        sheet.cell(row_idx, 2).fill = fill
        sheet.cell(row_idx, 1).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=label == "How to use")
        if label.startswith("Estimated base"):
            sheet.cell(row_idx, 2).number_format = "#,##0"
    sheet.merge_cells("B2:G2")
    sheet.row_dimensions[2].height = 120
    if all_ages:
        age_start_row = 1
        age_end_row = len(AGE_ORDER)
        age_list_range = f"'{AGE_OPTIONS_SHEET}'!$A${age_start_row}:$A${age_end_row}"
        workbook.defined_names.add(
            DefinedName(CITY_AGE_LIST_NAME, attr_text=age_list_range)
        )
        age_dv = DataValidation(type="list", formula1=age_list_range, allow_blank=False)
        age_dv.showDropDown = False
        age_dv.showInputMessage = True
        age_dv.showErrorMessage = True
        age_dv.errorStyle = "stop"
        age_dv.promptTitle = "Select Your City Age"
        age_dv.prompt = "Pick the age to use for ranking data."
        age_dv.errorTitle = "Use the dropdown"
        age_dv.error = "Choose a city age from the dropdown list."
        sheet.add_data_validation(age_dv)
        age_dv.add(CITY_AGE_CELL)
    if category_options:
        category_start_row = 1
        category_end_row = len(category_options)
        category_list_range = f"'{CATEGORY_OPTIONS_SHEET}'!$A${category_start_row}:$A${category_end_row}"
        workbook.defined_names.add(
            DefinedName(BUILDING_CATEGORY_LIST_NAME, attr_text=category_list_range)
        )
        category_dv = DataValidation(type="list", formula1=category_list_range, allow_blank=False)
        category_dv.showDropDown = False
        category_dv.showInputMessage = True
        category_dv.showErrorMessage = True
        category_dv.errorStyle = "stop"
        category_dv.promptTitle = "Building source category filter"
        category_dv.prompt = "Pick a category to show in the ranking sheets."
        category_dv.errorTitle = "Use the dropdown"
        category_dv.error = "Choose a building category from the dropdown list."
        sheet.add_data_validation(category_dv)
        category_dv.add(BUILDING_CATEGORY_FILTER_CELL)

    def add_focus_selector(
        row_idx: int,
        title: str,
        input_cell_ref: str,
        default_value: int,
        slider_labels: Sequence[str],
    ) -> None:
        sheet.cell(row_idx, 1, title)
        sheet.cell(row_idx, 1).font = Font(size=12, color=HEADER_FONT_COLOR)
        sheet.cell(row_idx, 1).border = border
        sheet.cell(row_idx, 1).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_idx, 2, "Selected scale")
        sheet.cell(row_idx, 2).font = Font(size=12, color=HEADER_FONT_COLOR)
        sheet.cell(row_idx, 2).fill = slider_fill
        sheet.cell(row_idx, 2).border = border
        sheet.cell(row_idx, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet[input_cell_ref] = default_value
        sheet[input_cell_ref].font = Font(bold=True, size=12, color=TITLE_FONT_COLOR)
        sheet[input_cell_ref].fill = editable_fill
        sheet[input_cell_ref].border = border
        sheet[input_cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        sheet[input_cell_ref].number_format = "0"
        track_labels = [slider_labels[0], "", slider_labels[2], "", slider_labels[4]]
        for offset, label in enumerate(track_labels, start=3):
            label_cell = sheet.cell(row_idx, offset, label)
            value_cell = sheet.cell(row_idx + 1, offset, offset - 2)
            label_cell.font = Font(size=12, color=TITLE_FONT_COLOR)
            label_cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            value_cell.border = Border(
                left=track_side if offset == 3 else no_side,
                right=track_side if offset == 7 else no_side,
                top=track_side,
                bottom=track_side,
            )
            value_cell.fill = slider_fill
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.font = Font(bold=True, size=12, color=TITLE_FONT_COLOR)
            scale_value = offset - 2
            sheet.conditional_formatting.add(
                f"{get_column_letter(offset)}{row_idx + 1}:{get_column_letter(offset)}{row_idx + 1}",
                FormulaRule(formula=[f"{input_cell_ref}={scale_value}"], fill=slider_selected_fill),
            )

    add_focus_selector(
        12,
        "Fighting GBG/GE focus",
        FIGHTING_GBG_GE_FOCUS_CELL,
        DEFAULT_FIGHTING_GBG_GE_FOCUS,
        ["GBG only", "Mostly GBG", "Half and half", "Mostly GE", "GE only"],
    )
    add_focus_selector(
        14,
        "Fighting Red/Blue focus",
        FIGHTING_RED_BLUE_FOCUS_CELL,
        DEFAULT_FIGHTING_RED_BLUE_FOCUS,
        ["Red only", "Mostly Red", "Half and half", "Mostly Blue", "Blue only"],
    )
    add_focus_selector(
        16,
        "Fighting Attack/Defense focus",
        FIGHTING_ATTACK_DEFENSE_FOCUS_CELL,
        DEFAULT_FIGHTING_ATTACK_DEFENSE_FOCUS,
        ["Attack only", "Mostly attack", "Half and half", "Mostly defense", "Defense only"],
    )
    add_focus_selector(
        18,
        "Fighting Current/Next Age unit focus",
        FIGHTING_UNIT_AGE_FOCUS_CELL,
        DEFAULT_FIGHTING_UNIT_AGE_FOCUS,
        ["Current only", "Mostly current", "Half and half", "Mostly next", "Next only"],
    )
    add_focus_selector(
        20,
        "Production FP/Goods focus",
        PRODUCTION_FP_GOODS_FOCUS_CELL,
        DEFAULT_PRODUCTION_FP_GOODS_FOCUS,
        ["FP only", "Mostly FP", "Half and half", "Mostly goods", "Goods only"],
    )
    focus_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=False)
    sheet.add_data_validation(focus_dv)
    focus_dv.add(FIGHTING_GBG_GE_FOCUS_CELL)
    focus_dv.add(FIGHTING_RED_BLUE_FOCUS_CELL)
    focus_dv.add(FIGHTING_UNIT_AGE_FOCUS_CELL)
    focus_dv.add(FIGHTING_ATTACK_DEFENSE_FOCUS_CELL)
    focus_dv.add(PRODUCTION_FP_GOODS_FOCUS_CELL)

    role_dv = DataValidation(type="list", formula1='"Both,Blue,Red"', allow_blank=False)
    sheet.add_data_validation(role_dv)
    role_dv.add(QI_FIGHTER_ROLE_CELL)

    sheet.row_dimensions[10].height = 24
    for row_idx in (12, 14, 16, 18, 20):
        sheet.row_dimensions[row_idx].height = 36
        sheet.row_dimensions[row_idx + 1].height = 28

    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 20
    for column in ("C", "D", "E", "F", "G"):
        sheet.column_dimensions[column].width = 16


def write_advanced_controls_sheet(
    workbook: Workbook,
    reference_file: str,
    era: str,
    attr_keys: Sequence[str],
    stats: Dict[str, Dict[str, float]],
    available_only: bool,
    record_count: int,
    all_ages: bool = False,
) -> None:
    sheet = workbook.create_sheet(ADVANCED_CONTROLS_SHEET)
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    editable_fill = PatternFill("solid", fgColor=EDITABLE_FILL_COLOR)
    context_fill = PatternFill("solid", fgColor=CONTROL_CONTEXT_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"] = "Advanced Building Attribute Ranking Controls"
    sheet["A1"].font = Font(bold=True, size=16, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=QI_WEIGHT_OVERRIDE_COLUMN)

    rows = [
        ("Selected city age" if all_ages else "Assumed age", f"={CONTROLS_SHEET_REF}!{CITY_AGE_CELL}" if all_ages else era),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("How to use", f"Optional fine tuning: keep Weight mode as Default to use generated weights, or switch to Custom and enter yellow override values in the right-side override columns. Set Weight mode back to Default to restore all generated defaults. After changes, review {OVERALL_RANKING_SHEET} or the focused ranking sheets."),
    ]
    for idx, (label, value) in enumerate(rows, start=2):
        sheet.cell(idx, 1, label)
        sheet.cell(idx, 2, value)
        sheet.cell(idx, 1).font = Font(bold=True)
        sheet.cell(idx, 2).fill = (
            context_fill
            if "age" in label.lower()
            else PatternFill(fill_type=None)
        )
        sheet.cell(idx, 2).alignment = Alignment(vertical="top", wrap_text=label == "How to use")
    sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=QI_WEIGHT_OVERRIDE_COLUMN)
    sheet["B4"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet.row_dimensions[4].height = 45
    sheet["A5"] = "Total active weight"
    sheet["A5"].font = Font(bold=True)
    total_headers = (
        (OVERALL_TOTAL_WEIGHT_CELL, "Overall"),
        (FIGHTING_TOTAL_WEIGHT_CELL, "Fighting"),
        (FP_GOODS_TOTAL_WEIGHT_CELL, "Farming"),
        (QI_TOTAL_WEIGHT_CELL, "QI"),
    )
    for total_cell, header in total_headers:
        col_idx = sheet[total_cell].column
        cell = sheet.cell(5, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right")
    sheet["A7"] = "Weight mode"
    sheet["B7"] = "Default"
    sheet["C7"] = "Default restores all generated weights. Custom uses nonblank values from the override columns."
    sheet["A7"].font = Font(bold=True)
    sheet[ADVANCED_WEIGHT_MODE_CELL].fill = editable_fill
    sheet[ADVANCED_WEIGHT_MODE_CELL].border = border
    sheet[ADVANCED_WEIGHT_MODE_CELL].alignment = Alignment(vertical="top", wrap_text=True)
    sheet["C7"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet.merge_cells(start_row=7, start_column=3, end_row=7, end_column=QI_WEIGHT_OVERRIDE_COLUMN)
    sheet[OVERALL_TOTAL_WEIGHT_CELL] = f"=SUM(D{WEIGHT_START_ROW}:D1048576)"
    sheet[FIGHTING_TOTAL_WEIGHT_CELL] = f"=SUM(F{WEIGHT_START_ROW}:F1048576)"
    sheet[FP_GOODS_TOTAL_WEIGHT_CELL] = f"=SUM(H{WEIGHT_START_ROW}:H1048576)"
    sheet[QI_TOTAL_WEIGHT_CELL] = f"=SUM(J{WEIGHT_START_ROW}:J1048576)"
    sheet[OVERALL_TOTAL_WEIGHT_CELL].font = Font(bold=True)
    sheet[FIGHTING_TOTAL_WEIGHT_CELL].font = Font(bold=True)
    sheet[FP_GOODS_TOTAL_WEIGHT_CELL].font = Font(bold=True)
    sheet[QI_TOTAL_WEIGHT_CELL].font = Font(bold=True)

    mode_dv = DataValidation(type="list", formula1='"Default,Custom"', allow_blank=False)
    mode_dv.showDropDown = False
    mode_dv.showInputMessage = True
    mode_dv.showErrorMessage = True
    mode_dv.errorStyle = "stop"
    mode_dv.promptTitle = "Weight mode"
    mode_dv.prompt = "Default restores generated weights. Custom uses nonblank override cells."
    mode_dv.errorTitle = "Use Default or Custom"
    mode_dv.error = "Choose Default or Custom from the dropdown."
    sheet.add_data_validation(mode_dv)
    mode_dv.add(ADVANCED_WEIGHT_MODE_CELL)

    header_row = WEIGHT_HEADER_ROW
    headers = [
        "Attribute",
        "Description",
        "Overall Weight",
        "Overall Abs Weight",
        "Fighting Weight",
        "Fighting Abs Weight",
        "Farming Weight",
        "Farming Abs Weight",
        "QI Weight",
        "QI Abs Weight",
        "Direction",
        "Min",
        "Max",
        "Overall Raw Weight",
        "Overall Weight Group",
        "Overall Weight Budget",
        "Overall Override",
        "Fighting Override",
        "Farming Override",
        "QI Override",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, key in enumerate(attr_keys, start=WEIGHT_START_ROW):
        attr_idx = row_idx - WEIGHT_START_ROW
        raw_col = get_column_letter(RAW_START_COLUMN + attr_idx)
        data_end = BUILDING_DATA_START_ROW + record_count - 1
        label = attr_label(key)
        force_zero_weight = is_forced_zero_weight_attr(key)
        sheet.cell(row_idx, 1, label)
        sheet.cell(row_idx, 2, attr_description(key))
        sheet.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(
            row_idx,
            3,
            0 if force_zero_weight else weight_mode_formula(overall_weight_cell_value(row_idx), OVERALL_WEIGHT_OVERRIDE_COLUMN, row_idx),
        )
        sheet.cell(row_idx, 4, f"=ABS(C{row_idx})")
        sheet.cell(
            row_idx,
            5,
            0 if force_zero_weight else weight_mode_formula(fighting_weight_cell_value(key), FIGHTING_WEIGHT_OVERRIDE_COLUMN, row_idx),
        )
        sheet.cell(row_idx, 6, f"=ABS(E{row_idx})")
        sheet.cell(
            row_idx,
            7,
            0 if force_zero_weight else weight_mode_formula(fp_goods_weight_cell_value(key), FP_GOODS_WEIGHT_OVERRIDE_COLUMN, row_idx),
        )
        sheet.cell(row_idx, 8, f"=ABS(G{row_idx})")
        sheet.cell(
            row_idx,
            9,
            0 if force_zero_weight else weight_mode_formula(qi_role_weight_formula(key) or qi_weight_for_attr(key), QI_WEIGHT_OVERRIDE_COLUMN, row_idx),
        )
        sheet.cell(row_idx, 10, f"=ABS(I{row_idx})")
        sheet.cell(row_idx, 11, direction_for_attr(key))
        if (all_ages or key in {PROD_FP_ATTR, PROD_GOODS_ATTR, PROD_GUILD_GOODS_ATTR, PROD_MEDALS_ATTR}) and record_count:
            sheet.cell(row_idx, 12, f"=MIN('{OVERALL_SOURCE_SHEET}'!${raw_col}${BUILDING_DATA_START_ROW}:${raw_col}${data_end})")
            max_formula = f"MAX('{OVERALL_SOURCE_SHEET}'!${raw_col}${BUILDING_DATA_START_ROW}:${raw_col}${data_end})"
            max_anchor_formula = normalization_max_anchor_formula_for_attr(key)
            if max_anchor_formula:
                max_formula = f"MAX({max_formula},{max_anchor_formula})"
            sheet.cell(row_idx, 13, f"={max_formula}")
        else:
            sheet.cell(row_idx, 12, numeric_cell(stats[key]["min"]))
            sheet.cell(row_idx, 13, numeric_cell(stats[key]["max"]))
        sheet.cell(row_idx, OVERALL_RAW_WEIGHT_COLUMN, 0 if force_zero_weight else overall_raw_weight_cell_value(key))
        weight_group = overall_weight_group_for_attr(key)
        sheet.cell(row_idx, OVERALL_WEIGHT_GROUP_COLUMN, weight_group)
        sheet.cell(row_idx, OVERALL_WEIGHT_BUDGET_COLUMN, overall_weight_budget_cell_value(weight_group))
        for col_idx in range(1, QI_WEIGHT_OVERRIDE_COLUMN + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.border = border
            if col_idx == 2:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in (
                OVERALL_WEIGHT_OVERRIDE_COLUMN,
                FIGHTING_WEIGHT_OVERRIDE_COLUMN,
                FP_GOODS_WEIGHT_OVERRIDE_COLUMN,
                QI_WEIGHT_OVERRIDE_COLUMN,
            ):
                cell.fill = editable_fill
            if col_idx in (
                3,
                4,
                5,
                6,
                7,
                8,
                9,
                10,
                12,
                13,
                OVERALL_RAW_WEIGHT_COLUMN,
                OVERALL_WEIGHT_BUDGET_COLUMN,
                OVERALL_WEIGHT_OVERRIDE_COLUMN,
                FIGHTING_WEIGHT_OVERRIDE_COLUMN,
                FP_GOODS_WEIGHT_OVERRIDE_COLUMN,
                QI_WEIGHT_OVERRIDE_COLUMN,
            ):
                cell.number_format = "0.00"
        if force_zero_weight:
            sheet.row_dimensions[row_idx].hidden = True

    for row_idx in range(WEIGHT_START_ROW, WEIGHT_HEADER_ROW + len(attr_keys) + 1):
        sheet.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=True)

    dv = DataValidation(type="decimal", operator="between", formula1="-1000", formula2="1000")
    sheet.add_data_validation(dv)
    if attr_keys:
        dv.add(f"{get_column_letter(OVERALL_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_START_ROW}:{get_column_letter(OVERALL_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_HEADER_ROW + len(attr_keys)}")
        dv.add(f"{get_column_letter(FIGHTING_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_START_ROW}:{get_column_letter(FIGHTING_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_HEADER_ROW + len(attr_keys)}")
        dv.add(f"{get_column_letter(FP_GOODS_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_START_ROW}:{get_column_letter(FP_GOODS_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_HEADER_ROW + len(attr_keys)}")
        dv.add(f"{get_column_letter(QI_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_START_ROW}:{get_column_letter(QI_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_HEADER_ROW + len(attr_keys)}")
    sheet.freeze_panes = f"A{WEIGHT_START_ROW}"
    sheet.auto_filter.ref = f"A{WEIGHT_HEADER_ROW}:{get_column_letter(QI_WEIGHT_OVERRIDE_COLUMN)}{WEIGHT_HEADER_ROW + len(attr_keys)}"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 92
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 12
    sheet.column_dimensions["I"].width = 12
    sheet.column_dimensions["J"].width = 12
    sheet.column_dimensions["K"].width = 12
    sheet.column_dimensions["L"].width = 14
    sheet.column_dimensions["M"].width = 14
    sheet.column_dimensions[get_column_letter(OVERALL_RAW_WEIGHT_COLUMN)].width = 16
    sheet.column_dimensions[get_column_letter(OVERALL_WEIGHT_GROUP_COLUMN)].width = 18
    sheet.column_dimensions[get_column_letter(OVERALL_WEIGHT_BUDGET_COLUMN)].width = 18
    sheet.column_dimensions[get_column_letter(OVERALL_WEIGHT_OVERRIDE_COLUMN)].width = 16
    sheet.column_dimensions[get_column_letter(FIGHTING_WEIGHT_OVERRIDE_COLUMN)].width = 16
    sheet.column_dimensions[get_column_letter(FP_GOODS_WEIGHT_OVERRIDE_COLUMN)].width = 16
    sheet.column_dimensions[get_column_letter(QI_WEIGHT_OVERRIDE_COLUMN)].width = 16
    sheet.column_dimensions["D"].hidden = True
    sheet.column_dimensions["F"].hidden = True
    sheet.column_dimensions["H"].hidden = True
    sheet.column_dimensions["J"].hidden = True
    sheet.column_dimensions[get_column_letter(OVERALL_RAW_WEIGHT_COLUMN)].hidden = True
    sheet.column_dimensions[get_column_letter(OVERALL_WEIGHT_GROUP_COLUMN)].hidden = True
    sheet.column_dimensions[get_column_letter(OVERALL_WEIGHT_BUDGET_COLUMN)].hidden = True
    for row in sheet.iter_rows(min_row=2, max_row=7, min_col=1, max_col=QI_WEIGHT_OVERRIDE_COLUMN):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet["B4"].alignment = Alignment(vertical="top", wrap_text=True)
    for total_cell, _header in total_headers:
        sheet.cell(5, sheet[total_cell].column).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)


def write_age_data_sheet(
    workbook: Workbook,
    records_by_age: Dict[str, List[Dict[str, Any]]],
    attr_keys: Sequence[str],
) -> Dict[str, Any]:
    sheet = workbook.create_sheet(AGE_DATA_SHEET)
    sheet.sheet_state = "hidden"
    headers = [
        "Lookup Key",
        "Entity ID",
        "Selected Age",
        "Size",
        REQUIRE_ROAD_HEADER,
        "Area",
        "Available By Age",
        "Environment Effect",
        "Fragment / Reward Production",
    ] + [attr_label(key) for key in attr_keys]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for age in AGE_ORDER:
        for record in records_by_age.get(age, []):
            base_values = [
                age_lookup_key(str(record["entity_id"]), age),
                record["entity_id"],
                age_display_name(age),
                record["size"],
                require_road_connection_label(record),
                numeric_cell(record["area"]) if record["area"] is not None else "",
                record["available"],
                record["environment_effect"],
                record["reward_production"],
            ]
            for col_idx, value in enumerate(base_values, start=1):
                sheet.cell(row_idx, col_idx, value)
            for attr_idx, key in enumerate(attr_keys, start=10):
                sheet.cell(row_idx, attr_idx, numeric_cell(float(record["attrs"].get(key, 0.0))))
            row_idx += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_idx - 1, 1)}"
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 10
    sheet.column_dimensions["G"].width = 17
    sheet.column_dimensions["H"].width = 48
    sheet.column_dimensions["I"].width = 92
    for col_idx in range(10, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    return {
        "max_row": max(row_idx - 1, 2),
        "size_col": "D",
        "road_col": "E",
        "area_col": "F",
        "available_col": "G",
        "environment_col": "H",
        "reward_col": "I",
        "attr_columns": {key: get_column_letter(10 + idx) for idx, key in enumerate(attr_keys)},
    }


def write_age_options_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(AGE_OPTIONS_SHEET)
    sheet.sheet_state = "hidden"
    for row_idx, age in enumerate(AGE_ORDER, start=1):
        sheet.cell(row_idx, 1, age_display_name(age))
    sheet.column_dimensions["A"].width = 28


def write_category_options_sheet(workbook: Workbook, category_options: Sequence[str]) -> None:
    sheet = workbook.create_sheet(CATEGORY_OPTIONS_SHEET)
    sheet.sheet_state = "hidden"
    for row_idx, category in enumerate(category_options, start=1):
        sheet.cell(row_idx, 1, category)
    sheet.column_dimensions["A"].width = 36


def write_goods_resource_audit_sheet(workbook: Workbook, attr_keys: Sequence[str]) -> None:
    sheet = workbook.create_sheet(GOODS_RESOURCE_AUDIT_SHEET)
    sheet.sheet_state = "hidden"
    sheet.sheet_view.showGridLines = False
    headers = ["Resource Key", "Classification", "Discovered In", "Notes"]
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row_values in enumerate(goods_resource_audit_rows(attr_keys), start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [32, 34, 24, 72]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def write_buildings_sheet(
    workbook: Workbook,
    records: List[Dict[str, Any]],
    attr_keys: Sequence[str],
    stats: Dict[str, Dict[str, float]],
    age_data_context: Optional[Dict[str, Any]] = None,
) -> None:
    sheet = workbook.create_sheet(OVERALL_SOURCE_SHEET)
    sheet.sheet_state = "hidden"
    sheet.sheet_view.showGridLines = False

    base_headers = [
        "Building",
        "Overall Rank",
        "Overall Score",
        "Size",
        REQUIRE_ROAD_HEADER,
        "Area",
    ]
    metadata_headers = [
        "Type",
        "Selected Age",
        "Available By Age",
        "Building Category",
        "Environment Effect",
        "Entity ID",
        "Fragment / Reward Production",
    ]
    raw_headers = [attr_label(key) for key in attr_keys]

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"] = OVERALL_RANKING_SHEET
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(base_headers) + len(raw_headers) + len(metadata_headers))

    coefficient_row = 2
    offset_row = 3
    fighting_coefficient_row = 4
    fighting_offset_row = 5
    fp_goods_coefficient_row = 6
    fp_goods_offset_row = 7
    qi_coefficient_row = 8
    qi_offset_row = 9
    header_row = BUILDING_HEADER_ROW
    all_headers = base_headers + raw_headers + metadata_headers
    for col_idx, header in enumerate(all_headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    raw_start = len(base_headers) + 1
    raw_end = raw_start + len(attr_keys) - 1
    attr_columns = {key: raw_start + idx for idx, key in enumerate(attr_keys)}
    metadata_start = raw_start + len(attr_keys)
    data_start = header_row + 1
    data_end = data_start + len(records) - 1

    def coefficient_formula_for(weight_col: str, controls_row: int, signed_centered: bool) -> str:
        if signed_centered:
            return (
                f'=IF(MAX(ABS({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}),'
                f'ABS({ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row}))=0,0,'
                f'{ADVANCED_CONTROLS_SHEET_REF}!${weight_col}${controls_row}*100/'
                f'MAX(ABS({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}),'
                f'ABS({ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row})))'
            )
        return (
            f'=IF({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}={ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row},0,'
            f'{ADVANCED_CONTROLS_SHEET_REF}!${weight_col}${controls_row}*IF({ADVANCED_CONTROLS_SHEET_REF}!$K${controls_row}="Lower",-1,1)'
            f'*100/({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}-{ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row}))'
        )

    def offset_formula_for(weight_col: str, controls_row: int, signed_centered: bool) -> str:
        if signed_centered:
            return "0"
        return (
            f'=IF({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}={ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row},0,'
            f'{ADVANCED_CONTROLS_SHEET_REF}!${weight_col}${controls_row}*IF({ADVANCED_CONTROLS_SHEET_REF}!$K${controls_row}="Lower",'
            f'{ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row},-{ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row})'
            f'*100/({ADVANCED_CONTROLS_SHEET_REF}!$M${controls_row}-{ADVANCED_CONTROLS_SHEET_REF}!$L${controls_row}))'
        )

    for idx, key in enumerate(attr_keys):
        controls_row = WEIGHT_START_ROW + idx
        signed_centered = key in SIGNED_CENTERED_ATTRS
        coefficient_formula = coefficient_formula_for("C", controls_row, signed_centered)
        offset_formula = offset_formula_for("C", controls_row, signed_centered)
        fighting_coefficient_formula = coefficient_formula_for("E", controls_row, signed_centered)
        fighting_offset_formula = offset_formula_for("E", controls_row, signed_centered)
        fp_goods_coefficient_formula = coefficient_formula_for("G", controls_row, signed_centered)
        fp_goods_offset_formula = offset_formula_for("G", controls_row, signed_centered)
        qi_coefficient_formula = coefficient_formula_for("I", controls_row, signed_centered)
        qi_offset_formula = offset_formula_for("I", controls_row, signed_centered)
        sheet.cell(coefficient_row, raw_start + idx, coefficient_formula)
        sheet.cell(offset_row, raw_start + idx, offset_formula)
        sheet.cell(fighting_coefficient_row, raw_start + idx, fighting_coefficient_formula)
        sheet.cell(fighting_offset_row, raw_start + idx, fighting_offset_formula)
        sheet.cell(fp_goods_coefficient_row, raw_start + idx, fp_goods_coefficient_formula)
        sheet.cell(fp_goods_offset_row, raw_start + idx, fp_goods_offset_formula)
        sheet.cell(qi_coefficient_row, raw_start + idx, qi_coefficient_formula)
        sheet.cell(qi_offset_row, raw_start + idx, qi_offset_formula)

    overall_weights = overall_weight_map(attr_keys)
    records.sort(
        key=lambda record: (
            -default_score(record, attr_keys, stats, lambda key: overall_weights.get(key, 0.0)),
            str(record["name"]),
        )
    )
    for row_idx, record in enumerate(records, start=data_start):
        entity_id = str(record["entity_id"])
        empty_text_formula = '""'
        values = [
            record["name"],
            None,
            None,
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['size_col'], age_data_context['max_row'], empty_text_formula)}"
                if age_data_context
                else record["size"]
            ),
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['road_col'], age_data_context['max_row'], empty_text_formula)}"
                if age_data_context
                else require_road_connection_label(record)
            ),
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['area_col'], age_data_context['max_row'], '0')}"
                if age_data_context
                else (numeric_cell(record["area"]) if record["area"] is not None else "")
            ),
        ]
        for col_idx, value in enumerate(values, start=1):
            sheet.cell(row_idx, col_idx, value)
        for attr_idx, key in enumerate(attr_keys):
            raw_value = float(record["attrs"].get(key, 0.0))
            if age_data_context:
                raw_formula = age_data_lookup_formula(
                    entity_id,
                    age_data_context["attr_columns"][key],
                    age_data_context["max_row"],
                    "0",
                )
                raw_expr = raw_formula
            else:
                raw_expr = cached_number(raw_value)
            cell = sheet.cell(row_idx, raw_start + attr_idx)
            if key == PROD_FP_ATTR and BOOST_FP_ATTR in attr_columns:
                boost_cell = f"{get_column_letter(attr_columns[BOOST_FP_ATTR])}{row_idx}"
                cell.value = f"={raw_expr}+{boost_cell}*{CONTROLS_SHEET_REF}!{ESTIMATED_FP_PRODUCTION_CELL}/100"
            elif key == PROD_GOODS_ATTR and (
                BOOST_GOODS_ATTR in attr_columns or BOOST_SPECIAL_GOODS_ATTR in attr_columns
            ):
                boost_terms = []
                if BOOST_GOODS_ATTR in attr_columns:
                    boost_terms.append(
                        f"{get_column_letter(attr_columns[BOOST_GOODS_ATTR])}{row_idx}"
                        f"*{CONTROLS_SHEET_REF}!{ESTIMATED_GOODS_PRODUCTION_CELL}/100"
                    )
                if BOOST_SPECIAL_GOODS_ATTR in attr_columns:
                    boost_terms.append(
                        f"{get_column_letter(attr_columns[BOOST_SPECIAL_GOODS_ATTR])}{row_idx}"
                        f"*{CONTROLS_SHEET_REF}!{ESTIMATED_SPECIAL_GOODS_PRODUCTION_CELL}/100"
                    )
                boost_expr = "+".join(boost_terms)
                cell.value = f"={raw_expr}+{boost_expr}"
            elif key == PROD_GUILD_GOODS_ATTR and BOOST_GUILD_GOODS_ATTR in attr_columns:
                boost_cell = f"{get_column_letter(attr_columns[BOOST_GUILD_GOODS_ATTR])}{row_idx}"
                cell.value = f"={raw_expr}+{boost_cell}*{CONTROLS_SHEET_REF}!{ESTIMATED_GUILD_GOODS_PRODUCTION_CELL}/100"
            elif key == PROD_MEDALS_ATTR and BOOST_MEDALS_ATTR in attr_columns:
                boost_cell = f"{get_column_letter(attr_columns[BOOST_MEDALS_ATTR])}{row_idx}"
                cell.value = f"={raw_expr}+{boost_cell}*{CONTROLS_SHEET_REF}!{ESTIMATED_MEDAL_PRODUCTION_CELL}/100"
            elif age_data_context:
                cell.value = f"={raw_expr}"
            else:
                cell.value = numeric_cell(raw_value)

        metadata_values = [
            record["type"],
            f"={CONTROLS_SHEET_REF}!{CITY_AGE_CELL}" if age_data_context else record["selected_age"],
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['available_col'], age_data_context['max_row'], empty_text_formula)}"
                if age_data_context
                else record["available"]
            ),
            building_category_label(str(record["entity_id"])),
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['environment_col'], age_data_context['max_row'], empty_text_formula)}"
                if age_data_context
                else record["environment_effect"]
            ),
            record["entity_id"],
            (
                f"={age_data_lookup_formula(entity_id, age_data_context['reward_col'], age_data_context['max_row'], empty_text_formula)}"
                if age_data_context
                else record["reward_production"]
            ),
        ]
        for col_idx, value in enumerate(metadata_values, start=metadata_start):
            sheet.cell(row_idx, col_idx, value)

        if attr_keys:
            raw_range = f"{get_column_letter(raw_start)}{row_idx}:{get_column_letter(raw_end)}{row_idx}"
            coefficient_range = f"${get_column_letter(raw_start)}${coefficient_row}:${get_column_letter(raw_end)}${coefficient_row}"
            offset_range = f"${get_column_letter(raw_start)}${offset_row}:${get_column_letter(raw_end)}${offset_row}"
            category_cell = f"{get_column_letter(metadata_start + 3)}{row_idx}"
            score_formula = (
                f"IF({ADVANCED_CONTROLS_SHEET_REF}!{OVERALL_TOTAL_WEIGHT_CELL}=0,0,"
                f"(SUMPRODUCT({raw_range},{coefficient_range})+SUM({offset_range}))/{ADVANCED_CONTROLS_SHEET_REF}!{OVERALL_TOTAL_WEIGHT_CELL})"
            )
            sheet.cell(
                row_idx,
                3,
                f"=IF(OR({CONTROLS_SHEET_REF}!{BUILDING_CATEGORY_FILTER_CELL}={excel_string(ALL_BUILDING_CATEGORIES)},{category_cell}={CONTROLS_SHEET_REF}!{BUILDING_CATEGORY_FILTER_CELL}),{score_formula},\"\")",
            )
        else:
            sheet.cell(row_idx, 3, 0)
        sheet.cell(row_idx, 2, f'=IF(C{row_idx}="","",1+COUNTIF($C${data_start}:$C${data_end},">"&C{row_idx}))')

    for row in sheet.iter_rows(min_row=data_start, max_row=data_end, min_col=1, max_col=len(all_headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == metadata_start + 6)
            if cell.column in (2, 3) or raw_start <= cell.column <= raw_end:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(all_headers))}{data_end}"
    sheet.row_dimensions[coefficient_row].hidden = True
    sheet.row_dimensions[offset_row].hidden = True
    sheet.row_dimensions[fighting_coefficient_row].hidden = True
    sheet.row_dimensions[fighting_offset_row].hidden = True
    sheet.row_dimensions[fp_goods_coefficient_row].hidden = True
    sheet.row_dimensions[fp_goods_offset_row].hidden = True
    sheet.row_dimensions[qi_coefficient_row].hidden = True
    sheet.row_dimensions[qi_offset_row].hidden = True

    widths = {
        "A": 34,
        "B": 10,
        "C": 12,
        "D": 10,
        "E": 20,
        "F": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for col_idx in range(raw_start, raw_end + 1):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 18
        if is_road_connection_attr_key(attr_keys[col_idx - raw_start]):
            sheet.column_dimensions[column_letter].hidden = True
    metadata_widths = [18, 18, 17, 24, 48, 28, 92]
    for offset, width in enumerate(metadata_widths):
        sheet.column_dimensions[get_column_letter(metadata_start + offset)].width = width
    sheet.column_dimensions[get_column_letter(metadata_start)].hidden = True
    sheet.column_dimensions[get_column_letter(metadata_start + 2)].hidden = True
    sheet.column_dimensions[get_column_letter(metadata_start + 5)].hidden = True
    apply_building_name_color_rules(
        sheet,
        data_start,
        data_end,
        metadata_start + 5,
        event_reward_abbreviations(records),
    )

    if records:
        score_range = f"C{data_start}:C{data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def write_overall_ranking_view_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(OVERALL_RANKING_SHEET)
    sheet.sheet_view.showGridLines = False

    base_headers = [
        "Building",
        "Overall Rank",
        "Overall Score",
        "Size",
        REQUIRE_ROAD_HEADER,
        "Area",
    ]
    metadata_headers = [
        "Type",
        "Selected Age",
        "Available By Age",
        "Building Category",
        "Environment Effect",
        "Entity ID",
        "Fragment / Reward Production",
        "Source Row",
    ]
    display_attr_keys = overall_ranking_display_attr_keys(attr_keys)
    raw_headers = [overall_ranking_attr_label(key) for key in display_attr_keys]
    all_headers = base_headers + raw_headers + metadata_headers

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_count = min(OVERALL_TOP_N, len(records))
    sheet["A1"] = f"Top {row_count} Overall Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))

    header_row = BUILDING_HEADER_ROW
    data_start = header_row + 1
    data_end = data_start + row_count - 1
    raw_start = len(base_headers) + 1
    raw_end = raw_start + len(display_attr_keys) - 1
    metadata_start = raw_start + len(display_attr_keys)
    source_metadata_start = RAW_START_COLUMN + len(attr_keys)
    source_row_col = metadata_start + len(metadata_headers) - 1
    source_row_cell_col = get_column_letter(source_row_col)

    for col_idx, header in enumerate(all_headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    score_range = f"'{OVERALL_SCORE_SHEET}'!$B${score_data_start}:$B${score_data_end}"
    match_range = f"'{OVERALL_SCORE_SHEET}'!$H${score_data_start}:$H${score_data_end}"
    source_row_range = f"'{OVERALL_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"
    source_mapping = {
        1: "A",
        4: "D",
        5: "E",
        6: "F",
        metadata_start: get_column_letter(source_metadata_start),
        metadata_start + 1: get_column_letter(source_metadata_start + 1),
        metadata_start + 2: get_column_letter(source_metadata_start + 2),
        metadata_start + 3: get_column_letter(source_metadata_start + 3),
        metadata_start + 4: get_column_letter(source_metadata_start + 4),
        metadata_start + 5: get_column_letter(source_metadata_start + 5),
        metadata_start + 6: get_column_letter(source_metadata_start + 6),
    }

    for row_idx in range(data_start, data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"{source_row_cell_col}{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE(FILTER({score_range},{match_range}=1),{relative_rank}),"")')
        sheet.cell(
            row_idx,
            source_row_col,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},({score_range}={score_cell})*({match_range}=1)),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            if output_col == metadata_start + 6:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell})&"")')
            else:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(display_attr_keys, start=raw_start):
            if key == NET_HAPPINESS_ATTR:
                gross_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness"))
                demand_expr = "0"
                if "happiness_demanded" in attr_keys:
                    demand_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness_demanded"))
                    demand_expr = f"INDEX('{OVERALL_SOURCE_SHEET}'!${demand_col}:${demand_col},${source_row_cell})"
                sheet.cell(
                    row_idx,
                    attr_idx,
                    f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${gross_col}:${gross_col},${source_row_cell})+{demand_expr})',
                )
            else:
                source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
                sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    for row in sheet.iter_rows(min_row=data_start, max_row=data_end, min_col=1, max_col=len(all_headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == metadata_start + 6)
            if cell.column in (2, 3, 6) or raw_start <= cell.column <= raw_end:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(all_headers))}{data_end}"
    for row_idx in range(2, header_row):
        sheet.row_dimensions[row_idx].hidden = True
    widths = {
        "A": 34,
        "B": 10,
        "C": 12,
        "D": 10,
        "E": 20,
        "F": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for col_idx in range(raw_start, raw_end + 1):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 18
        if is_road_connection_attr_key(display_attr_keys[col_idx - raw_start]):
            sheet.column_dimensions[column_letter].hidden = True
    metadata_widths = [18, 18, 17, 24, 48, 28, 92, 10]
    for offset, width in enumerate(metadata_widths):
        sheet.column_dimensions[get_column_letter(metadata_start + offset)].width = width
    sheet.column_dimensions[get_column_letter(metadata_start)].hidden = True
    sheet.column_dimensions[get_column_letter(metadata_start + 2)].hidden = True
    sheet.column_dimensions[get_column_letter(metadata_start + 5)].hidden = True
    sheet.column_dimensions[get_column_letter(source_row_col)].hidden = True
    apply_building_name_color_rules(
        sheet,
        data_start,
        data_end,
        metadata_start + 5,
        event_reward_abbreviations(records),
    )

    if records:
        sheet.conditional_formatting.add(
            f"C{data_start}:C{data_end}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def write_overall_scores_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(OVERALL_SCORE_SHEET)
    sheet.sheet_state = "hidden"

    headers = [
        "Source Row",
        "Overall Score",
        "Overall Score Rank",
        "Adjusted Area",
        "Overall Efficiency Score",
        "Overall Efficiency Rank",
        "Building Category",
        "Category Match",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 2
    data_end = data_start + len(records) - 1
    raw_start = RAW_START_COLUMN
    raw_end = raw_start + len(attr_keys) - 1
    all_buildings_raw_start = get_column_letter(raw_start)
    all_buildings_raw_end = get_column_letter(raw_end)
    score_range = f"$B${data_start}:$B${data_end}"
    efficiency_score_range = f"$E${data_start}:$E${data_end}"
    match_range = f"$H${data_start}:$H${data_end}"
    for idx, record in enumerate(records):
        row_idx = data_start + idx
        source_row = BUILDING_DATA_START_ROW + idx
        sheet.cell(row_idx, 1, source_row)
        if attr_keys:
            raw_range = f"'{OVERALL_SOURCE_SHEET}'!{all_buildings_raw_start}{source_row}:{all_buildings_raw_end}{source_row}"
            coefficient_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$2:${all_buildings_raw_end}$2"
            offset_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$3:${all_buildings_raw_end}$3"
            sheet.cell(
                row_idx,
                2,
                f"=IF({ADVANCED_CONTROLS_SHEET_REF}!{OVERALL_TOTAL_WEIGHT_CELL}=0,0,(SUMPRODUCT({raw_range},{coefficient_range})+SUM({offset_range}))/{ADVANCED_CONTROLS_SHEET_REF}!{OVERALL_TOTAL_WEIGHT_CELL})",
            )
        else:
            sheet.cell(row_idx, 2, 0)
        sheet.cell(row_idx, 3, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({score_range}>B{row_idx})*({match_range}=1)))')
        area_formula = f"'{OVERALL_SOURCE_SHEET}'!$F${source_row}+IF('{OVERALL_SOURCE_SHEET}'!$E${source_row}=\"Y\",1,0)"
        sheet.cell(row_idx, 4, f"={area_formula}")
        sheet.cell(row_idx, 5, f"=IF(D{row_idx}=0,0,B{row_idx}/D{row_idx})")
        sheet.cell(row_idx, 6, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({efficiency_score_range}>E{row_idx})*({match_range}=1)))')
        sheet.cell(row_idx, 7, building_category_label(str(record["entity_id"])))
        sheet.cell(row_idx, 8, building_category_match_formula(f"G{row_idx}"))

    for col_idx in range(1, 9):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 16


def write_fighting_scores_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(FIGHTING_SCORE_SHEET)
    sheet.sheet_state = "hidden"

    headers = [
        "Source Row",
        "Fighting Score",
        "Fighting Score Rank",
        "Adjusted Area",
        "Fighting Efficiency Score",
        "Fighting Efficiency Rank",
        "Building Category",
        "Category Match",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 2
    data_end = data_start + len(records) - 1
    raw_start = RAW_START_COLUMN
    raw_end = raw_start + len(attr_keys) - 1
    all_buildings_raw_start = get_column_letter(raw_start)
    all_buildings_raw_end = get_column_letter(raw_end)
    score_range = f"$B${data_start}:$B${data_end}"
    efficiency_score_range = f"$E${data_start}:$E${data_end}"
    match_range = f"$H${data_start}:$H${data_end}"
    for idx, record in enumerate(records):
        row_idx = data_start + idx
        source_row = BUILDING_DATA_START_ROW + idx
        sheet.cell(row_idx, 1, source_row)
        if attr_keys:
            raw_range = f"'{OVERALL_SOURCE_SHEET}'!{all_buildings_raw_start}{source_row}:{all_buildings_raw_end}{source_row}"
            coefficient_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$4:${all_buildings_raw_end}$4"
            offset_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$5:${all_buildings_raw_end}$5"
            sheet.cell(
                row_idx,
                2,
                f"=IF({ADVANCED_CONTROLS_SHEET_REF}!{FIGHTING_TOTAL_WEIGHT_CELL}=0,0,(SUMPRODUCT({raw_range},{coefficient_range})+SUM({offset_range}))/{ADVANCED_CONTROLS_SHEET_REF}!{FIGHTING_TOTAL_WEIGHT_CELL})",
            )
        else:
            sheet.cell(row_idx, 2, 0)
        sheet.cell(row_idx, 3, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({score_range}>B{row_idx})*({match_range}=1)))')
        area_formula = f"'{OVERALL_SOURCE_SHEET}'!$F${source_row}+IF('{OVERALL_SOURCE_SHEET}'!$E${source_row}=\"Y\",1,0)"
        sheet.cell(row_idx, 4, f"={area_formula}")
        sheet.cell(row_idx, 5, f"=IF(D{row_idx}=0,0,B{row_idx}/D{row_idx})")
        sheet.cell(row_idx, 6, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({efficiency_score_range}>E{row_idx})*({match_range}=1)))')
        sheet.cell(row_idx, 7, building_category_label(str(record["entity_id"])))
        sheet.cell(row_idx, 8, building_category_match_formula(f"G{row_idx}"))

    for col_idx in range(1, 9):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 16


def write_fighting_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet("Fighting Ranking")
    sheet.sheet_view.showGridLines = False

    fighting_attr_keys = [key for key in attr_keys if fighting_weight_for_attr(key)]
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "Fighting Rank",
        "Fighting Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in fighting_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} Fighting Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Uses the Fighting Weight controls. Non-fighting attributes default to zero in that control set."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    data_end_all = BUILDING_DATA_START_ROW + len(records) - 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    fighting_score_range = f"'{FIGHTING_SCORE_SHEET}'!$B${score_data_start}:$B${score_data_end}"
    fighting_source_row_range = f"'{FIGHTING_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        4: "D",
        5: "E",
        6: "F",
        7: "G",
        8: "H",
        9: "J",
        10: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"K{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({fighting_score_range},{relative_rank}),"")')
        sheet.cell(
            row_idx,
            11,
            f'=IF({score_cell}="","",INDEX(FILTER({fighting_source_row_range},{fighting_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(fighting_attr_keys, start=12):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (9,))
            if cell.column in (2, 3, 8) or cell.column >= 12:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 14,
        "D": 18,
        "E": 18,
        "F": 17,
        "G": 10,
        "H": 10,
        "I": 92,
        "J": 28,
        "K": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["K"].hidden = True
    for col_idx in range(12, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_fighting_efficiency_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(FIGHTING_EFFICIENCY_SHEET)
    sheet.sheet_view.showGridLines = False

    fighting_attr_keys = [key for key in attr_keys if fighting_weight_for_attr(key)]
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "Efficiency Rank",
        "Fighting Efficiency Score",
        "Fighting Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Adjusted Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in fighting_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} Fighting Efficiency Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Fighting efficiency is Fighting Score divided by adjusted area. Adjusted area adds 1 when the building requires a road connection."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    efficiency_score_range = f"'{FIGHTING_SCORE_SHEET}'!$E${score_data_start}:$E${score_data_end}"
    fighting_source_row_range = f"'{FIGHTING_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        5: "D",
        6: "E",
        7: "F",
        8: "G",
        9: "H",
        11: "J",
        12: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"M{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({efficiency_score_range},{relative_rank}),"")')
        sheet.cell(row_idx, 4, f'=IF(${source_row_cell}="","",INDEX(\'{FIGHTING_SCORE_SHEET}\'!$B:$B,MATCH(${source_row_cell},\'{FIGHTING_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(row_idx, 10, f'=IF(${source_row_cell}="","",INDEX(\'{FIGHTING_SCORE_SHEET}\'!$D:$D,MATCH(${source_row_cell},\'{FIGHTING_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(
            row_idx,
            13,
            f'=IF({score_cell}="","",INDEX(FILTER({fighting_source_row_range},{efficiency_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(fighting_attr_keys, start=14):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (11,))
            if cell.column in (2, 3, 4, 9, 10) or cell.column >= 14:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 18,
        "D": 14,
        "E": 18,
        "F": 18,
        "G": 17,
        "H": 10,
        "I": 10,
        "J": 13,
        "K": 92,
        "L": 28,
        "M": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["M"].hidden = True
    for col_idx in range(14, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def fp_goods_display_attr_keys(attr_keys: Sequence[str]) -> List[str]:
    preferred = [
        PROD_FP_ATTR,
        PROD_GOODS_ATTR,
        PROD_MEDALS_ATTR,
        NET_HAPPINESS_ATTR,
        "prod_resource_blueprint",
        "prod_resource_premium",
        "prod_resource_supplies",
        "prod_resource_all_goods_of_previous_age",
        "prod_resource_all_goods_of_age",
        "prod_resource_all_goods_of_next_age",
        "prod_resource_special_goods_up_to_age",
        PROD_GUILD_GOODS_ATTR,
        BOOST_FP_ATTR,
        BOOST_GOODS_ATTR,
        BOOST_SPECIAL_GOODS_ATTR,
    ]
    out = [key for key in preferred if key in attr_keys]
    for key in attr_keys:
        label = attr_label(key)
        if key not in out and (
            fp_goods_weight_for_attr(key)
            or label in {"Production: FPs", "Production: Goods Total", "Production: Guild Goods"}
            or "Goods" in label
        ):
            out.append(key)
    return out


def qi_display_attr_keys(attr_keys: Sequence[str]) -> List[str]:
    return [key for key in attr_keys if is_qi_attr(key)]


def street_connection_column(attr_keys: Sequence[str]) -> Optional[str]:
    for key in attr_keys:
        if is_road_connection_attr_key(key):
            return get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
    return None


def write_ranked_score_sheet(
    workbook: Workbook,
    sheet_name: str,
    score_sheet_name: str,
    rank_header: str,
    score_header: str,
    title: str,
    note: str,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
    display_attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    display_attr_keys = [key for key in display_attr_keys if not is_road_connection_attr_key(key)]

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    attr_start = 7
    metadata_start = attr_start + len(display_attr_keys)
    type_col = metadata_start
    selected_age_col = metadata_start + 1
    available_col = metadata_start + 2
    entity_col = metadata_start + 3
    source_row_col = metadata_start + 4
    fragment_col = metadata_start + 5
    all_buildings_metadata_start = RAW_START_COLUMN + len(attr_keys)
    headers = (
        ["Building", rank_header, score_header, "Size", REQUIRE_ROAD_HEADER, "Area"]
        + [attr_label(key) for key in display_attr_keys]
        + [
            "Type",
            "Selected Age",
            "Available By Age",
            "Entity ID",
            "Source Row",
            "Fragment / Reward Production",
        ]
    )

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = note
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    score_range = f"'{score_sheet_name}'!$B${score_data_start}:$B${score_data_end}"
    match_range = f"'{score_sheet_name}'!$H${score_data_start}:$H${score_data_end}"
    source_row_range = f"'{score_sheet_name}'!$A${score_data_start}:$A${score_data_end}"
    source_row_cell_col = get_column_letter(source_row_col)

    source_mapping = {
        1: "A",
        4: "D",
        5: "E",
        6: "F",
        type_col: get_column_letter(all_buildings_metadata_start),
        selected_age_col: get_column_letter(all_buildings_metadata_start + 1),
        available_col: get_column_letter(all_buildings_metadata_start + 2),
        entity_col: get_column_letter(all_buildings_metadata_start + 5),
        fragment_col: get_column_letter(all_buildings_metadata_start + 6),
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"{source_row_cell_col}{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE(FILTER({score_range},{match_range}=1),{relative_rank}),"")')
        sheet.cell(
            row_idx,
            source_row_col,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},({score_range}={score_cell})*({match_range}=1)),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            if output_col == fragment_col:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell})&"")')
            else:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(display_attr_keys, start=attr_start):
            if key == NET_HAPPINESS_ATTR:
                gross_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness"))
                demand_expr = "0"
                if "happiness_demanded" in attr_keys:
                    demand_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness_demanded"))
                    demand_expr = f"INDEX('{OVERALL_SOURCE_SHEET}'!${demand_col}:${demand_col},${source_row_cell})"
                sheet.cell(
                    row_idx,
                    attr_idx,
                    f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${gross_col}:${gross_col},${source_row_cell})+{demand_expr})',
                )
            else:
                source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
                sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == fragment_col)
            if cell.column in (2, 3, 6) or attr_start <= cell.column < metadata_start:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_data_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_data_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        1: 34,
        2: 12,
        3: 14,
        4: 10,
        5: 20,
        6: 10,
        type_col: 18,
        selected_age_col: 18,
        available_col: 17,
        entity_col: 28,
        source_row_col: 10,
        fragment_col: 92,
    }
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.column_dimensions[source_row_cell_col].hidden = True
    sheet.column_dimensions[get_column_letter(type_col)].hidden = True
    sheet.column_dimensions[get_column_letter(available_col)].hidden = True
    sheet.column_dimensions[get_column_letter(entity_col)].hidden = True
    for col_idx in range(attr_start, metadata_start):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    apply_building_name_color_rules(
        sheet,
        data_start,
        output_data_end,
        entity_col,
        event_reward_abbreviations(records),
    )


def write_ranked_efficiency_sheet(
    workbook: Workbook,
    sheet_name: str,
    score_sheet_name: str,
    efficiency_score_header: str,
    base_score_header: str,
    title: str,
    note: str,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
    display_attr_keys: Sequence[str],
    top_n: int = FIGHTING_TOP_N,
    attr_label_func: Callable[[str], str] = attr_label,
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    display_attr_keys = [key for key in display_attr_keys if not is_road_connection_attr_key(key)]

    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    attr_start = 9
    metadata_start = attr_start + len(display_attr_keys)
    type_col = metadata_start
    selected_age_col = metadata_start + 1
    available_col = metadata_start + 2
    entity_col = metadata_start + 3
    source_row_col = metadata_start + 4
    fragment_col = metadata_start + 5
    all_buildings_metadata_start = RAW_START_COLUMN + len(attr_keys)
    headers = (
        [
            "Building",
            "Efficiency Rank",
            efficiency_score_header,
            base_score_header,
            "Size",
            REQUIRE_ROAD_HEADER,
            "Area",
            "Adjusted Area",
        ]
        + [attr_label_func(key) for key in display_attr_keys]
        + [
            "Type",
            "Selected Age",
            "Available By Age",
            "Entity ID",
            "Source Row",
            "Fragment / Reward Production",
        ]
    )

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = note
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(top_n, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    efficiency_score_range = f"'{score_sheet_name}'!$E${score_data_start}:$E${score_data_end}"
    match_range = f"'{score_sheet_name}'!$H${score_data_start}:$H${score_data_end}"
    source_row_range = f"'{score_sheet_name}'!$A${score_data_start}:$A${score_data_end}"
    source_row_cell_col = get_column_letter(source_row_col)

    source_mapping = {
        1: "A",
        5: "D",
        6: "E",
        7: "F",
        type_col: get_column_letter(all_buildings_metadata_start),
        selected_age_col: get_column_letter(all_buildings_metadata_start + 1),
        available_col: get_column_letter(all_buildings_metadata_start + 2),
        entity_col: get_column_letter(all_buildings_metadata_start + 5),
        fragment_col: get_column_letter(all_buildings_metadata_start + 6),
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"{source_row_cell_col}{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE(FILTER({efficiency_score_range},{match_range}=1),{relative_rank}),"")')
        sheet.cell(row_idx, 4, f'=IF(${source_row_cell}="","",INDEX(\'{score_sheet_name}\'!$B:$B,MATCH(${source_row_cell},\'{score_sheet_name}\'!$A:$A,0)))')
        sheet.cell(row_idx, 8, f'=IF(${source_row_cell}="","",INDEX(\'{score_sheet_name}\'!$D:$D,MATCH(${source_row_cell},\'{score_sheet_name}\'!$A:$A,0)))')
        sheet.cell(
            row_idx,
            source_row_col,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},({efficiency_score_range}={score_cell})*({match_range}=1)),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            if output_col == fragment_col:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell})&"")')
            else:
                sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(display_attr_keys, start=attr_start):
            if key == NET_HAPPINESS_ATTR:
                gross_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness"))
                demand_expr = "0"
                if "happiness_demanded" in attr_keys:
                    demand_col = get_column_letter(RAW_START_COLUMN + attr_keys.index("happiness_demanded"))
                    demand_expr = f"INDEX('{OVERALL_SOURCE_SHEET}'!${demand_col}:${demand_col},${source_row_cell})"
                sheet.cell(
                    row_idx,
                    attr_idx,
                    f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${gross_col}:${gross_col},${source_row_cell})+{demand_expr})',
                )
            else:
                source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
                sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == fragment_col)
            if cell.column in (2, 3, 4, 7, 8) or attr_start <= cell.column < metadata_start:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_data_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_data_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        1: 34,
        2: 12,
        3: 18,
        4: 14,
        5: 10,
        6: 20,
        7: 10,
        8: 13,
        type_col: 18,
        selected_age_col: 18,
        available_col: 17,
        entity_col: 28,
        source_row_col: 10,
        fragment_col: 92,
    }
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.column_dimensions[source_row_cell_col].hidden = True
    sheet.column_dimensions[get_column_letter(type_col)].hidden = True
    sheet.column_dimensions[get_column_letter(available_col)].hidden = True
    sheet.column_dimensions[get_column_letter(entity_col)].hidden = True
    for col_idx in range(attr_start, metadata_start):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    apply_building_name_color_rules(
        sheet,
        data_start,
        output_data_end,
        entity_col,
        event_reward_abbreviations(records),
    )


def write_fp_goods_scores_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(FP_GOODS_SCORE_SHEET)
    sheet.sheet_state = "hidden"

    headers = [
        "Source Row",
        "Farming Score",
        "Farming Score Rank",
        "Adjusted Area",
        "Farming Efficiency Score",
        "Farming Efficiency Rank",
        "Building Category",
        "Category Match",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 2
    data_end = data_start + len(records) - 1
    raw_start = RAW_START_COLUMN
    raw_end = raw_start + len(attr_keys) - 1
    all_buildings_raw_start = get_column_letter(raw_start)
    all_buildings_raw_end = get_column_letter(raw_end)
    score_range = f"$B${data_start}:$B${data_end}"
    efficiency_score_range = f"$E${data_start}:$E${data_end}"
    match_range = f"$H${data_start}:$H${data_end}"
    for idx, record in enumerate(records):
        row_idx = data_start + idx
        source_row = BUILDING_DATA_START_ROW + idx
        sheet.cell(row_idx, 1, source_row)
        if attr_keys:
            raw_range = f"'{OVERALL_SOURCE_SHEET}'!{all_buildings_raw_start}{source_row}:{all_buildings_raw_end}{source_row}"
            coefficient_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$6:${all_buildings_raw_end}$6"
            offset_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$7:${all_buildings_raw_end}$7"
            sheet.cell(
                row_idx,
                2,
                f"=IF({ADVANCED_CONTROLS_SHEET_REF}!{FP_GOODS_TOTAL_WEIGHT_CELL}=0,0,(SUMPRODUCT({raw_range},{coefficient_range})+SUM({offset_range}))/{ADVANCED_CONTROLS_SHEET_REF}!{FP_GOODS_TOTAL_WEIGHT_CELL})",
            )
        else:
            sheet.cell(row_idx, 2, 0)
        sheet.cell(row_idx, 3, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({score_range}>B{row_idx})*({match_range}=1)))')
        area_formula = f"'{OVERALL_SOURCE_SHEET}'!$F${source_row}+IF('{OVERALL_SOURCE_SHEET}'!$E${source_row}=\"Y\",1,0)"
        sheet.cell(row_idx, 4, f"={area_formula}")
        sheet.cell(row_idx, 5, f"=IF(D{row_idx}=0,0,B{row_idx}/D{row_idx})")
        sheet.cell(row_idx, 6, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({efficiency_score_range}>E{row_idx})*({match_range}=1)))')
        sheet.cell(row_idx, 7, building_category_label(str(record["entity_id"])))
        sheet.cell(row_idx, 8, building_category_match_formula(f"G{row_idx}"))

    for col_idx in range(1, 9):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 16


def write_fp_goods_ranking_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(FP_GOODS_PRODUCTION_SHEET)
    sheet.sheet_view.showGridLines = False

    production_attr_keys = fp_goods_display_attr_keys(attr_keys)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "Farming Rank",
        "Farming Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in production_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} Farming Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Uses the Farming Weight controls. Default primary weight is split between FPs and Goods Total; secondary weight covers medals, net happiness, blueprints, diamonds, and supplies."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    fp_goods_score_range = f"'{FP_GOODS_SCORE_SHEET}'!$B${score_data_start}:$B${score_data_end}"
    source_row_range = f"'{FP_GOODS_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        4: "D",
        5: "E",
        6: "F",
        7: "G",
        8: "H",
        9: "J",
        10: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"K{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({fp_goods_score_range},{relative_rank}),"")')
        sheet.cell(
            row_idx,
            11,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},{fp_goods_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(production_attr_keys, start=12):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (9,))
            if cell.column in (2, 3, 8) or cell.column >= 12:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 14,
        "D": 18,
        "E": 18,
        "F": 17,
        "G": 10,
        "H": 10,
        "I": 92,
        "J": 28,
        "K": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["K"].hidden = True
    for col_idx in range(12, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_fp_goods_efficiency_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(FP_GOODS_EFFICIENCY_SHEET)
    sheet.sheet_view.showGridLines = False

    production_attr_keys = fp_goods_display_attr_keys(attr_keys)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "Efficiency Rank",
        "Farming Efficiency Score",
        "Farming Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Adjusted Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in production_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} Farming Efficiency Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Farming efficiency is Farming Score divided by adjusted area. Default farming weight includes FPs, Goods Total, medals, net happiness, blueprints, diamonds, and supplies."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    efficiency_score_range = f"'{FP_GOODS_SCORE_SHEET}'!$E${score_data_start}:$E${score_data_end}"
    source_row_range = f"'{FP_GOODS_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        5: "D",
        6: "E",
        7: "F",
        8: "G",
        9: "H",
        11: "J",
        12: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"M{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({efficiency_score_range},{relative_rank}),"")')
        sheet.cell(row_idx, 4, f'=IF(${source_row_cell}="","",INDEX(\'{FP_GOODS_SCORE_SHEET}\'!$B:$B,MATCH(${source_row_cell},\'{FP_GOODS_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(row_idx, 10, f'=IF(${source_row_cell}="","",INDEX(\'{FP_GOODS_SCORE_SHEET}\'!$D:$D,MATCH(${source_row_cell},\'{FP_GOODS_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(
            row_idx,
            13,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},{efficiency_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(production_attr_keys, start=14):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (11,))
            if cell.column in (2, 3, 4, 9, 10) or cell.column >= 14:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 18,
        "D": 14,
        "E": 18,
        "F": 18,
        "G": 17,
        "H": 10,
        "I": 10,
        "J": 13,
        "K": 92,
        "L": 28,
        "M": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["M"].hidden = True
    for col_idx in range(14, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_qi_scores_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(QI_SCORE_SHEET)
    sheet.sheet_state = "hidden"

    headers = [
        "Source Row",
        "QI Score",
        "QI Score Rank",
        "Adjusted Area",
        "QI Efficiency Score",
        "QI Efficiency Rank",
        "Building Category",
        "Category Match",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx, header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 2
    data_end = data_start + len(records) - 1
    raw_start = RAW_START_COLUMN
    raw_end = raw_start + len(attr_keys) - 1
    all_buildings_raw_start = get_column_letter(raw_start)
    all_buildings_raw_end = get_column_letter(raw_end)
    score_range = f"$B${data_start}:$B${data_end}"
    efficiency_score_range = f"$E${data_start}:$E${data_end}"
    match_range = f"$H${data_start}:$H${data_end}"
    for idx, record in enumerate(records):
        row_idx = data_start + idx
        source_row = BUILDING_DATA_START_ROW + idx
        sheet.cell(row_idx, 1, source_row)
        if attr_keys:
            raw_range = f"'{OVERALL_SOURCE_SHEET}'!{all_buildings_raw_start}{source_row}:{all_buildings_raw_end}{source_row}"
            coefficient_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$8:${all_buildings_raw_end}$8"
            offset_range = f"'{OVERALL_SOURCE_SHEET}'!${all_buildings_raw_start}$9:${all_buildings_raw_end}$9"
            sheet.cell(
                row_idx,
                2,
                f"=IF({ADVANCED_CONTROLS_SHEET_REF}!{QI_TOTAL_WEIGHT_CELL}=0,0,(SUMPRODUCT({raw_range},{coefficient_range})+SUM({offset_range}))/{ADVANCED_CONTROLS_SHEET_REF}!{QI_TOTAL_WEIGHT_CELL})",
            )
        else:
            sheet.cell(row_idx, 2, 0)
        sheet.cell(row_idx, 3, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({score_range}>B{row_idx})*({match_range}=1)))')
        area_formula = f"'{OVERALL_SOURCE_SHEET}'!$F${source_row}+IF('{OVERALL_SOURCE_SHEET}'!$E${source_row}=\"Y\",1,0)"
        sheet.cell(row_idx, 4, f"={area_formula}")
        sheet.cell(row_idx, 5, f"=IF(D{row_idx}=0,0,B{row_idx}/D{row_idx})")
        sheet.cell(row_idx, 6, f'=IF(H{row_idx}=0,"",1+SUMPRODUCT(({efficiency_score_range}>E{row_idx})*({match_range}=1)))')
        sheet.cell(row_idx, 7, building_category_label(str(record["entity_id"])))
        sheet.cell(row_idx, 8, building_category_match_formula(f"G{row_idx}"))

    for col_idx in range(1, 9):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 16


def write_qi_ranking_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(QI_RANKING_SHEET)
    sheet.sheet_view.showGridLines = False

    qi_attr_keys = qi_display_attr_keys(attr_keys)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "QI Rank",
        "QI Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in qi_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} QI Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Uses the QI Weight controls. Includes attributes whose data key or label is QI-related."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    qi_score_range = f"'{QI_SCORE_SHEET}'!$B${score_data_start}:$B${score_data_end}"
    source_row_range = f"'{QI_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        4: "D",
        5: "E",
        6: "F",
        7: "G",
        8: "H",
        9: "J",
        10: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"K{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({qi_score_range},{relative_rank}),"")')
        sheet.cell(
            row_idx,
            11,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},{qi_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(qi_attr_keys, start=12):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (9,))
            if cell.column in (2, 3, 8) or cell.column >= 12:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 14,
        "D": 18,
        "E": 18,
        "F": 17,
        "G": 10,
        "H": 10,
        "I": 92,
        "J": 28,
        "K": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["K"].hidden = True
    for col_idx in range(12, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_qi_efficiency_sheet(
    workbook: Workbook,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
) -> None:
    sheet = workbook.create_sheet(QI_EFFICIENCY_SHEET)
    sheet.sheet_view.showGridLines = False

    qi_attr_keys = qi_display_attr_keys(attr_keys)
    title_fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Building",
        "Efficiency Rank",
        "QI Efficiency Score",
        "QI Score",
        "Type",
        "Selected Age",
        "Available By Age",
        "Size",
        "Area",
        "Adjusted Area",
        "Fragment / Reward Production",
        "Entity ID",
        "Source Row",
    ] + [attr_label(key) for key in qi_attr_keys]

    sheet["A1"] = f"Top {min(FIGHTING_TOP_N, len(records))} QI Efficiency Buildings"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "QI efficiency is QI Score divided by adjusted area. Adjusted area adds 1 when the building requires a road connection."
    sheet["A2"].alignment = Alignment(wrap_text=False)

    header_row = 4
    data_start = header_row + 1
    row_count = min(FIGHTING_TOP_N, len(records))
    output_data_end = data_start + row_count - 1

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_data_start = 2
    score_data_end = score_data_start + len(records) - 1
    efficiency_score_range = f"'{QI_SCORE_SHEET}'!$E${score_data_start}:$E${score_data_end}"
    source_row_range = f"'{QI_SCORE_SHEET}'!$A${score_data_start}:$A${score_data_end}"

    source_mapping = {
        1: "A",
        5: "D",
        6: "E",
        7: "F",
        8: "G",
        9: "H",
        11: "J",
        12: "K",
    }

    for row_idx in range(data_start, output_data_end + 1):
        relative_rank = row_idx - data_start + 1
        score_cell = f"C{row_idx}"
        source_row_cell = f"M{row_idx}"
        sheet.cell(row_idx, 2, f'=IF({score_cell}="","",ROWS($B${data_start}:B{row_idx}))')
        sheet.cell(row_idx, 3, f'=IFERROR(LARGE({efficiency_score_range},{relative_rank}),"")')
        sheet.cell(row_idx, 4, f'=IF(${source_row_cell}="","",INDEX(\'{QI_SCORE_SHEET}\'!$B:$B,MATCH(${source_row_cell},\'{QI_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(row_idx, 10, f'=IF(${source_row_cell}="","",INDEX(\'{QI_SCORE_SHEET}\'!$D:$D,MATCH(${source_row_cell},\'{QI_SCORE_SHEET}\'!$A:$A,0)))')
        sheet.cell(
            row_idx,
            13,
            f'=IF({score_cell}="","",INDEX(FILTER({source_row_range},{efficiency_score_range}={score_cell}),COUNTIF($C${data_start}:{score_cell},{score_cell})))',
        )
        for output_col, source_col in source_mapping.items():
            sheet.cell(row_idx, output_col, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')
        for attr_idx, key in enumerate(qi_attr_keys, start=14):
            source_col = get_column_letter(RAW_START_COLUMN + attr_keys.index(key))
            sheet.cell(row_idx, attr_idx, f'=IF(${source_row_cell}="","",INDEX(\'{OVERALL_SOURCE_SHEET}\'!${source_col}:${source_col},${source_row_cell}))')

    max_col = len(headers)
    for row in sheet.iter_rows(min_row=data_start, max_row=max(output_data_end, header_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (11,))
            if cell.column in (2, 3, 4, 9, 10) or cell.column >= 14:
                cell.number_format = "0.00"

    sheet.freeze_panes = f"B{header_row + 1}"
    if row_count:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{output_data_end}"
        score_range = f"C{data_start}:C{output_data_end}"
        sheet.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 34,
        "B": 12,
        "C": 18,
        "D": 14,
        "E": 18,
        "F": 18,
        "G": 17,
        "H": 10,
        "I": 10,
        "J": 13,
        "K": 92,
        "L": 28,
        "M": 10,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.column_dimensions["M"].hidden = True
    for col_idx in range(14, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18


def write_about_sheet(
    workbook: Workbook,
    reference_file: str,
    era: str,
    records: Sequence[Dict[str, Any]],
    attr_keys: Sequence[str],
    available_only: bool,
    all_ages: bool = False,
) -> None:
    sheet = workbook.create_sheet("About")
    sheet.sheet_view.showGridLines = False
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet["A1"] = "Workbook Guide"
    sheet["A1"].font = Font(bold=True, size=15, color=TITLE_FONT_COLOR)
    sheet["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL_COLOR)
    sheet["A1"].border = border
    sheet.merge_cells("A1:B1")
    notes = [
        ("Reference file", display_path(reference_file)),
        ("Default selected age" if all_ages else "Assumed age", selected_age_display(era, all_ages)),
        ("Version", WORKBOOK_VERSION),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Buildings included", len(records)),
        ("Attributes discovered", len(attr_keys)),
        ("Available only filter", "Yes" if available_only else "No"),
        ("Estimated base FP production default", format_amount(DEFAULT_ESTIMATED_FP_PRODUCTION)),
        ("Estimated base regular goods production default", format_amount(DEFAULT_ESTIMATED_GOODS_PRODUCTION)),
        ("Estimated base special goods production default", format_amount(DEFAULT_ESTIMATED_SPECIAL_GOODS_PRODUCTION)),
        ("Estimated base guild goods production default", format_amount(DEFAULT_ESTIMATED_GUILD_GOODS_PRODUCTION)),
        ("Estimated base medal production default", format_amount(DEFAULT_ESTIMATED_MEDAL_PRODUCTION)),
        ("Start here", "Use Main Controls first. Pick your city age, enter your city's estimated base production for FP, regular goods, special goods, guild goods, and medals without percentage-based boosts, then choose the fighting, QI, and Production FP/Goods focus settings that match your priorities." if all_ages else "Use Main Controls first. Enter your city's estimated base production for FP, regular goods, special goods, guild goods, and medals without percentage-based boosts, then choose the fighting, QI, and Production FP/Goods focus settings that match your priorities."),
        ("All-age mode", "The selected city age on Main Controls updates the age-sensitive source values used by the ranking and efficiency sheets." if all_ages else "This workbook was generated for one fixed assumed age."),
        ("Scale controls", "Each 1-5 scale works left to right: 1 fully favors the left option, 3 is balanced, and 5 fully favors the right option."),
        ("Fighting focus", "Main Controls lets you tune GBG vs GE, red vs blue army use, attack vs defense boosts, and current-age vs next-age unit production."),
        ("Production FP/Goods focus", f"Main Controls cell {PRODUCTION_FP_GOODS_FOCUS_CELL.replace('$', '')} tunes FP vs goods production value. Default 2 is FP-heavy; 3 is balanced. The setting affects both Overall Ranking and Farming ranking weights."),
        ("QI role", "Choose whether QI fighting value should favor blue, red, or both roles."),
        ("Advanced controls", "Use Advanced Controls only for fine tuning. Leave Weight mode as Default to restore generated weights, or switch to Custom and enter yellow override values in the right-side override columns. A higher override weight makes that attribute matter more; zero turns it off."),
        ("Building source category filter", "Use the Main Controls source category dropdown to show all buildings or only a color-coded reward/event category on the ranking sheets."),
        ("Production boost conversion", "Boost percentages use matching Main Controls estimates: regular goods boost uses regular goods production, special goods boost uses special goods production, guild goods boost uses guild goods production, and FP/medal boosts use their own totals."),
        ("Base coin and supplies production", "Base Production: Coin and Base Production: Supplies on the Overall Ranking and Overall Efficiency Ranking sheets do not include percentage-based boosts."),
        ("Goods total", "Regular goods rollup: named goods, all/random goods of previous/current/next age, special goods up to age, and era_goods. It excludes FP, medals, money, supplies, guild goods, and settlement resources."),
        ("Guild goods", "Guild goods are tracked separately from regular goods, so changing one estimate does not change the other."),
        ("Happiness", "Gross Happiness shows happiness provided by the building. Net Happiness subtracts internal happiness demand; demand is still tracked as a negative scoring input but is not shown as a separate ranking column."),
        ("Overall ranking", f"Use {OVERALL_RANKING_SHEET} for a broad building comparison across production, fighting, and other weighted attributes. Overall fighting uses fixed sub-budgets for all non-QI combat, GBG, GE, QI, and unit production, and those sub-budgets follow the Main Controls fighting focus and QI fighter role settings."),
        ("Production normalization", "Major production scores are anchored to the Main Controls production assumptions so one extreme building does not define the full scoring range by itself."),
        ("Footprint", "Overall Ranking does not directly score footprint. Overall Efficiency divides Overall Score by adjusted area, adding one tile when a road connection is required."),
        ("Efficiency rankings", "Efficiency sheets favor buildings that score well for their footprint. Buildings that require a road connection are treated as needing one extra tile."),
        ("Fighting ranking", "Fighting Ranking uses your fighting focus settings and shows the top 100 fighting buildings."),
        ("Farming ranking", FARMING_RANKING_ABOUT_NOTE),
        ("QI ranking", "QI Ranking focuses on QI-related boosts, starting resources, action points, and the selected QI fighter role."),
        ("Building name colors", "Building names are color-coded by Entity ID: GBG rewards use W_MultiAge_GBG, QI rewards use W_MultiAge_GR, GE rewards use W_MultiAge_Expedition or W_MultiAge_GEX, and current-year event rewards use W_MultiAge_<event abbreviation><two-digit year><letter>... . Each workbook build scans the input data for current-year event abbreviations; newly detected events get an unused color, and assigned event colors stay fixed for the rest of that year."),
        ("Data scope", "Uses CityEntities reference definitions only; no placed-city quantities are used. Great buildings, QI settlement entities, and native era buildings are intentionally excluded from this reward-building workbook."),
    ]
    for row_idx, (label, value) in enumerate(notes, start=2):
        sheet.cell(row_idx, 1, label)
        sheet.cell(row_idx, 2, value)
        sheet.cell(row_idx, 1).font = Font(bold=True)
        sheet.cell(row_idx, 1).border = border
        sheet.cell(row_idx, 2).border = border
        sheet.cell(row_idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100


def apply_sheet_tab_colors(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        color = TAB_COLORS.get(sheet.title)
        if color:
            sheet.sheet_properties.tabColor = color


def build_workbook(reference_file: str, era: str, output_file: str, available_only: bool, all_ages: bool = False) -> None:
    payload = load_payload(reference_file)
    entities = payload.get("CityEntities")
    if not isinstance(entities, dict):
        raise SystemExit(f"CityEntities not found in reference file: {reference_file}")

    age_data_context: Optional[Dict[str, Any]] = None
    records_by_age: Dict[str, List[Dict[str, Any]]] = {}
    if all_ages:
        records_by_age, attr_keys = build_age_records(entities, list(AGE_ORDER), available_only)
        records = list(records_by_age.get(era, []))
        seen_entity_ids = {str(record["entity_id"]) for record in records}
        for age_records in records_by_age.values():
            for record in age_records:
                entity_id = str(record["entity_id"])
                if entity_id not in seen_entity_ids:
                    records.append(record)
                    seen_entity_ids.add(entity_id)
    else:
        records, attr_keys = collect_records(entities, era, available_only)
    stats = compute_attribute_stats(records, attr_keys)
    category_options = building_category_options(records)

    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.calcOnSave = True
    workbook.calculation.calcCompleted = False
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    write_controls_sheet(workbook, reference_file, era, available_only, all_ages, category_options)
    write_advanced_controls_sheet(workbook, reference_file, era, attr_keys, stats, available_only, len(records), all_ages)
    write_category_options_sheet(workbook, category_options)
    write_goods_resource_audit_sheet(workbook, attr_keys)
    if all_ages:
        write_age_options_sheet(workbook)
        age_data_context = write_age_data_sheet(workbook, records_by_age, attr_keys)
    write_buildings_sheet(workbook, records, attr_keys, stats, age_data_context)
    write_overall_scores_sheet(workbook, records, attr_keys)
    write_overall_ranking_view_sheet(workbook, records, attr_keys)
    write_ranked_efficiency_sheet(
        workbook,
        OVERALL_EFFICIENCY_SHEET,
        OVERALL_SCORE_SHEET,
        "Overall Efficiency Score",
        "Overall Score",
        f"Top {min(OVERALL_TOP_N, len(records))} Overall Efficiency Buildings",
        "Overall efficiency is Overall Score divided by adjusted area. Adjusted area adds 1 when the building requires a road connection.",
        records,
        attr_keys,
        overall_ranking_display_attr_keys(attr_keys),
        top_n=OVERALL_TOP_N,
        attr_label_func=overall_ranking_attr_label,
    )
    write_fighting_scores_sheet(workbook, records, attr_keys)
    write_ranked_score_sheet(
        workbook,
        "Fighting Ranking",
        FIGHTING_SCORE_SHEET,
        "Fighting Rank",
        "Fighting Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} Fighting Buildings",
        "Uses the Fighting Weight controls. Non-fighting attributes default to zero in that control set.",
        records,
        attr_keys,
        [key for key in attr_keys if fighting_weight_for_attr(key)],
    )
    write_ranked_efficiency_sheet(
        workbook,
        FIGHTING_EFFICIENCY_SHEET,
        FIGHTING_SCORE_SHEET,
        "Fighting Efficiency Score",
        "Fighting Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} Fighting Efficiency Buildings",
        "Fighting efficiency is Fighting Score divided by adjusted area. Adjusted area adds 1 when the building requires a road connection.",
        records,
        attr_keys,
        [key for key in attr_keys if fighting_weight_for_attr(key)],
    )
    write_fp_goods_scores_sheet(workbook, records, attr_keys)
    write_ranked_score_sheet(
        workbook,
        FP_GOODS_PRODUCTION_SHEET,
        FP_GOODS_SCORE_SHEET,
        "Farming Rank",
        "Farming Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} Farming Buildings",
        "Uses the Farming Weight controls. Default primary weight is split between FPs and Goods Total; secondary weight covers medals, net happiness, blueprints, diamonds, and supplies.",
        records,
        attr_keys,
        fp_goods_display_attr_keys(attr_keys),
    )
    write_ranked_efficiency_sheet(
        workbook,
        FP_GOODS_EFFICIENCY_SHEET,
        FP_GOODS_SCORE_SHEET,
        "Farming Efficiency Score",
        "Farming Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} Farming Efficiency Buildings",
        "Farming efficiency is Farming Score divided by adjusted area. Default farming weight includes FPs, Goods Total, medals, net happiness, blueprints, diamonds, and supplies.",
        records,
        attr_keys,
        fp_goods_display_attr_keys(attr_keys),
    )
    write_qi_scores_sheet(workbook, records, attr_keys)
    write_ranked_score_sheet(
        workbook,
        QI_RANKING_SHEET,
        QI_SCORE_SHEET,
        "QI Rank",
        "QI Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} QI Buildings",
        "Uses the QI Weight controls. Includes attributes whose data key or label is QI-related.",
        records,
        attr_keys,
        qi_display_attr_keys(attr_keys),
    )
    write_ranked_efficiency_sheet(
        workbook,
        QI_EFFICIENCY_SHEET,
        QI_SCORE_SHEET,
        "QI Efficiency Score",
        "QI Score",
        f"Top {min(FIGHTING_TOP_N, len(records))} QI Efficiency Buildings",
        "QI efficiency is QI Score divided by adjusted area. Adjusted area adds 1 when the building requires a road connection.",
        records,
        attr_keys,
        qi_display_attr_keys(attr_keys),
    )
    write_about_sheet(workbook, reference_file, era, records, attr_keys, available_only, all_ages)
    apply_sheet_tab_colors(workbook)

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    workbook.save(output_file)
    populate_formula_caches(output_file, records, attr_keys, stats)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--reference",
        default=DEFAULT_REFERENCE,
        help="Reference JSON file containing CityEntities. Defaults to input/ref/zpwd-ref.",
    )
    parser.add_argument(
        "--era",
        default="VirtualFuture",
        choices=list(AGE_ORDER),
        help="Assumed user age used for age-specific building attributes.",
    )
    parser.add_argument(
        "--available-only",
        action="store_true",
        help="Only include buildings whose native era is at or before the selected age.",
    )
    parser.add_argument(
        "--all_ages",
        "--all-ages",
        dest="all_ages",
        action="store_true",
        help="Build an all-age workbook with a Main Controls dropdown for the selected city age.",
    )
    parser.add_argument(
        "--output",
        help=(
            "Output .xlsx path. Defaults to output/building_attribute_ranking_<reference>_<era>.xlsx, "
            "or output/Interactive_Building_Rankings-All_Ages.xlsx in all-age mode."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    reference_file = os.path.abspath(args.reference)
    output_file = args.output
    if not output_file:
        if args.all_ages:
            output_file = os.path.join(OUTPUT_DIR, "Interactive_Building_Rankings-All_Ages.xlsx")
        else:
            token = safe_output_token(reference_file)
            output_file = os.path.join(OUTPUT_DIR, f"building_attribute_ranking_{token}_{args.era}.xlsx")
    build_workbook(reference_file, args.era, os.path.abspath(output_file), args.available_only, args.all_ages)
    print(f"Wrote {display_path(output_file)}")


if __name__ == "__main__":
    main()
