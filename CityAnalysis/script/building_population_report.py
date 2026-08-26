#!/usr/bin/env python3
"""Report placed city buildings that provide or consume population."""
from __future__ import annotations

import argparse
import json
import math
import os
from collections import Counter
from glob import glob
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
EXCLUDED_MAP_TYPES = {
    "friends_tavern",
    "greatbuilding",
    "hub_main",
    "hub_part",
    "main_building",
    "off_grid",
    "outpost_ship",
    "street",
}

AGE_BY_LEVEL = {
    0: "StoneAge",
    1: "BronzeAge",
    2: "IronAge",
    3: "EarlyMiddleAge",
    4: "HighMiddleAge",
    5: "LateMiddleAge",
    6: "ColonialAge",
    7: "IndustrialAge",
    8: "ProgressiveEra",
    9: "ModernEra",
    10: "PostModernEra",
    11: "ContemporaryEra",
    12: "TomorrowEra",
    13: "FutureEra",
    14: "ArcticFuture",
    15: "OceanicFuture",
    16: "VirtualFuture",
    17: "SpaceAgeMars",
    18: "SpaceAgeAsteroidBelt",
    19: "SpaceAgeVenus",
    20: "SpaceAgeJupiterMoon",
    21: "SpaceAgeTitan",
    22: "SpaceAgeSpaceHub",
    23: "StellarAgeDiscovery",
}


def latest_city_file() -> str:
    files = glob(os.path.join(INPUT_DIR, "city_*.json"))
    if not files:
        raise SystemExit(f"No city JSON files found in {INPUT_DIR}")
    return max(files, key=os.path.getmtime)


def load_payload(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        raw = json.load(handle)
    if isinstance(raw, dict) and isinstance(raw.get("data"), dict):
        return raw["data"]
    if isinstance(raw, dict):
        return raw
    raise SystemExit("Unexpected JSON payload format")


def extract_response_city_map(raw: Any) -> Optional[List[Dict[str, Any]]]:
    if not isinstance(raw, list):
        return None
    for item in raw:
        if not isinstance(item, dict):
            continue
        response_data = item.get("responseData")
        if not isinstance(response_data, dict):
            continue
        city_map = response_data.get("city_map")
        if not isinstance(city_map, dict):
            continue
        entities = city_map.get("entities")
        if isinstance(entities, list):
            return [entry for entry in entities if isinstance(entry, dict)]
    return None


def load_city_context(path: str) -> Tuple[Any, Dict[str, Any], str]:
    with open(path, "r", encoding="utf-8") as handle:
        raw = json.load(handle)

    response_city_map = extract_response_city_map(raw)
    if response_city_map is not None:
        reference_path = os.path.join(INPUT_DIR, "ref", "zpwd-ref")
        reference_payload = load_payload(reference_path)
        entities = reference_payload.get("CityEntities")
        if not isinstance(entities, dict):
            raise SystemExit(f"CityEntities not found in reference file: {reference_path}")
        return response_city_map, entities, reference_path

    if isinstance(raw, dict) and isinstance(raw.get("data"), dict):
        payload = raw["data"]
    elif isinstance(raw, dict):
        payload = raw
    else:
        raise SystemExit("Unexpected JSON payload format")

    city_map = payload.get("CityMapData")
    if not isinstance(city_map, (dict, list)):
        raise SystemExit("CityMapData not found in JSON")

    entities = payload.get("CityEntities")
    if not isinstance(entities, dict):
        raise SystemExit("CityEntities not found in JSON")

    return city_map, entities, path


def iter_map_items(city_map: Any) -> Iterable[Tuple[str, Dict[str, Any]]]:
    if isinstance(city_map, dict):
        for key, value in city_map.items():
            if isinstance(value, dict):
                yield str(key), value
        return
    if isinstance(city_map, list):
        for idx, value in enumerate(city_map):
            if isinstance(value, dict):
                yield str(idx), value


def detect_current_era(city_map: Any) -> Optional[str]:
    for _key, entry in iter_map_items(city_map):
        if entry.get("type") != "main_building":
            continue
        cityentity_id = entry.get("cityentity_id")
        if not isinstance(cityentity_id, str):
            continue
        parts = cityentity_id.split("_")
        if len(parts) >= 3 and parts[-1] == "Townhall":
            return parts[1]
    return None


def as_float(value: Any) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def population_from_resource_container(container: Any) -> Optional[float]:
    if isinstance(container, dict):
        value = as_float(container.get("population"))
        if value is not None:
            return value

        nested = population_from_resource_container(container.get("resources"))
        if nested is not None:
            return nested

        if container.get("type") == "population":
            for key in ("amount", "value"):
                typed_value = as_float(container.get(key))
                if typed_value is not None:
                    return typed_value
        return None

    if isinstance(container, list):
        total = 0.0
        found = False
        for item in container:
            value = population_from_resource_container(item)
            if value is not None:
                total += value
                found = True
        return total if found else None

    return None


def component_candidates(
    entity: Dict[str, Any],
    map_level: Optional[int],
    fallback_era: str,
) -> List[Tuple[str, Dict[str, Any]]]:
    components = entity.get("components")
    if not isinstance(components, dict):
        return []

    keys: List[str] = []
    if map_level in AGE_BY_LEVEL:
        keys.append(AGE_BY_LEVEL[map_level])
    keys.append(fallback_era)
    keys.append("AllAge")

    out: List[Tuple[str, Dict[str, Any]]] = []
    seen = set()
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        component = components.get(key)
        if isinstance(component, dict):
            out.append((key, component))
    return out


def era_keys(map_level: Optional[int], fallback_era: str) -> List[str]:
    keys: List[str] = []
    if map_level in AGE_BY_LEVEL:
        keys.append(AGE_BY_LEVEL[map_level])
    keys.append(fallback_era)

    out: List[str] = []
    seen = set()
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out


def entity_level_candidates(
    entity: Dict[str, Any],
    map_level: Optional[int],
    fallback_era: str,
) -> List[Tuple[str, Dict[str, Any]]]:
    entity_levels = entity.get("entity_levels")
    if not isinstance(entity_levels, list):
        return []

    wanted_eras = era_keys(map_level, fallback_era)
    out: List[Tuple[str, Dict[str, Any]]] = []
    seen = set()
    for wanted_era in wanted_eras:
        for entity_level in entity_levels:
            if not isinstance(entity_level, dict):
                continue
            if entity_level.get("era") != wanted_era:
                continue
            key = id(entity_level)
            if key in seen:
                continue
            seen.add(key)
            out.append((wanted_era, entity_level))
    return out


def extract_static_population(
    entity: Dict[str, Any],
    map_level: Optional[int],
    fallback_era: str,
) -> Tuple[float, str]:
    zero_source = ""
    for component_key, component in component_candidates(entity, map_level, fallback_era):
        value = population_from_resource_container(component.get("staticResources"))
        if value is not None:
            source = f"staticResources ({component_key})"
            if not math.isclose(value, 0.0):
                return value, source
            if not zero_source:
                zero_source = source

    for entity_level_key, entity_level in entity_level_candidates(entity, map_level, fallback_era):
        for field in ("population", "provided_population"):
            value = as_float(entity_level.get(field))
            if value is None:
                continue
            source = f"entity_levels.{field} ({entity_level_key})"
            if not math.isclose(value, 0.0):
                return value, source
            if not zero_source:
                zero_source = source

    value = population_from_resource_container(entity.get("staticResources"))
    if value is not None:
        return value, "staticResources"

    return 0.0, zero_source


def extract_required_population(
    entity: Dict[str, Any],
    map_level: Optional[int],
    fallback_era: str,
) -> Tuple[float, str]:
    zero_source = ""
    for component_key, component in component_candidates(entity, map_level, fallback_era):
        value = population_from_resource_container(component.get("requirements", {}).get("cost"))
        if value is not None:
            source = f"requirements.cost ({component_key})"
            if not math.isclose(value, 0.0):
                return value, source
            if not zero_source:
                zero_source = source

    for entity_level_key, entity_level in entity_level_candidates(entity, map_level, fallback_era):
        value = as_float(entity_level.get("required_population"))
        if value is not None:
            return value, f"entity_levels.required_population ({entity_level_key})"

    value = population_from_resource_container(entity.get("requirements", {}).get("cost"))
    if value is not None:
        return value, "requirements.cost"

    return 0.0, zero_source


def normalized_size(x_or_width: int, y_or_length: int) -> Tuple[int, int]:
    return y_or_length, x_or_width


def extract_size(
    entity: Dict[str, Any],
    map_level: Optional[int],
    fallback_era: str,
) -> Optional[Tuple[int, int]]:
    for _component_key, component in component_candidates(entity, map_level, fallback_era):
        placement = component.get("placement")
        if not isinstance(placement, dict):
            continue
        size = placement.get("size")
        if not isinstance(size, dict):
            continue
        x = size.get("x")
        y = size.get("y")
        if isinstance(x, int) and isinstance(y, int):
            return normalized_size(x, y)

    width = entity.get("width")
    length = entity.get("length")
    if isinstance(width, int) and isinstance(length, int):
        return normalized_size(width, length)
    return None


def format_number(value: float) -> str:
    if math.isclose(value, round(value)):
        return f"{int(round(value)):,}"
    return f"{value:,.2f}"


def display_path(path: str) -> str:
    abs_path = os.path.abspath(path)
    try:
        return os.path.relpath(abs_path, BASE_DIR)
    except ValueError:
        return os.path.basename(path)


def safe_output_token(path: str) -> str:
    name = os.path.basename(path.rstrip(os.sep))
    if not name:
        name = "city"
    if "." in name:
        name = os.path.splitext(name)[0]
    return "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in name)


def collect_population_records(
    city_map: Any,
    entities: Dict[str, Any],
    era: str,
) -> Tuple[List[Dict[str, Any]], int]:
    counts: Counter[Tuple[str, Optional[int], str]] = Counter()
    missing_defs = 0

    for _key, entry in iter_map_items(city_map):
        cityentity_id = entry.get("cityentity_id")
        if not isinstance(cityentity_id, str):
            continue
        level = entry.get("level")
        map_level = level if isinstance(level, int) else None
        entity_type = str(entry.get("type", ""))
        if entity_type in EXCLUDED_MAP_TYPES:
            continue
        if cityentity_id not in entities:
            missing_defs += 1
            continue
        counts[(cityentity_id, map_level, entity_type)] += 1

    records: List[Dict[str, Any]] = []
    for (cityentity_id, map_level, entity_type), count in counts.items():
        entity = entities.get(cityentity_id)
        if not isinstance(entity, dict):
            missing_defs += count
            continue

        static_pop, static_source = extract_static_population(entity, map_level, era)
        required_pop, required_source = extract_required_population(entity, map_level, era)
        effective_static_pop = static_pop - required_pop
        net_per_building = effective_static_pop
        size = extract_size(entity, map_level, era)
        area = None
        size_label = "unknown"
        if isinstance(size, tuple):
            area = size[0] * size[1]
            size_label = f"{size[0]}x{size[1]}"

        records.append(
            {
                "id": cityentity_id,
                "name": entity.get("name", cityentity_id),
                "type": entity.get("type", entity_type),
                "map_type": entity_type,
                "level": map_level,
                "count": count,
                "size_label": size_label,
                "area": area,
                "static_pop": effective_static_pop,
                "net_per_building": net_per_building,
                "total_net": net_per_building * count,
                "per_tile": net_per_building / area if area else 0.0,
                "static_source": static_source,
                "required_source": required_source,
            }
        )

    return records, missing_defs


def write_section(lines: List[str], title: str, records: List[Dict[str, Any]]) -> None:
    lines.append(title)
    lines.append("-" * len(title))
    if not records:
        lines.append("none")
        lines.append("")
        return

    for idx, info in enumerate(records, start=1):
        level = info.get("level")
        level_label = str(level) if level is not None else "n/a"
        per_tile_label = f"{info['per_tile']:,.2f}" if info.get("area") else "n/a"
        lines.append(
            f"{idx}. {info['name']} | count {info['count']} | level {level_label} | size {info['size_label']}"
        )
        lines.append(
            f"   Net population: {format_number(info['total_net'])} total | {format_number(info['net_per_building'])} each | {per_tile_label}/tile"
        )
        if info["static_pop"]:
            static_label = "Provides/static population"
            if info["static_pop"] < 0:
                static_label = "Consumes/static population"
            lines.append(f"   {static_label}: {format_number(abs(info['static_pop']))} each")
        source_parts = [part for part in (info.get("static_source"), info.get("required_source")) if part]
        if source_parts:
            lines.append(f"   Source: {', '.join(source_parts)}")
        lines.append(f"   CityEntity ID: {info['id']} | type {info['type']} | map type {info['map_type']}")
        lines.append("")


def write_report(
    path: str,
    source_file: str,
    reference_file: str,
    era: str,
    records: List[Dict[str, Any]],
    missing_defs: int,
) -> None:
    providers = [record for record in records if record["total_net"] > 0]
    consumers = [record for record in records if record["total_net"] < 0]
    neutral = [record for record in records if math.isclose(record["total_net"], 0.0)]

    providers.sort(key=lambda item: (-item["total_net"], -item["per_tile"], item["name"]))
    consumers.sort(key=lambda item: (item["total_net"], item["per_tile"], item["name"]))

    total_provided = sum(record["total_net"] for record in providers)
    total_consumed = -sum(record["total_net"] for record in consumers)
    total_net = total_provided - total_consumed
    placed_count = sum(record["count"] for record in records)

    lines: List[str] = []
    lines.append(f"Source file: {display_path(source_file)}")
    if reference_file != source_file:
        lines.append(f"Reference file: {display_path(reference_file)}")
    lines.append(f"Era: {era}")
    lines.append("Population convention: positive static population provides population; required population is shown as negative static population.")
    lines.append(f"Excluded map types: {', '.join(sorted(EXCLUDED_MAP_TYPES))}")
    lines.append(f"Placed building groups: {len(records)}")
    lines.append(f"Placed building count: {placed_count}")
    lines.append(f"Population provided: {format_number(total_provided)}")
    lines.append(f"Population consumed: {format_number(total_consumed)}")
    lines.append(f"Net population: {format_number(total_net)}")
    lines.append(f"Neutral groups omitted from rankings: {len(neutral)}")
    lines.append(f"Placed buildings missing CityEntities definitions: {missing_defs}")
    lines.append("")

    write_section(lines, "Population Providers", providers)
    write_section(lines, "Population Consumers", consumers)

    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines).rstrip() + "\n")


def split_records(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    providers = [record for record in records if record["total_net"] > 0]
    consumers = [record for record in records if record["total_net"] < 0]
    neutral = [record for record in records if math.isclose(record["total_net"], 0.0)]
    providers.sort(key=lambda item: (-item["total_net"], -item["per_tile"], item["name"]))
    consumers.sort(key=lambda item: (item["total_net"], item["per_tile"], item["name"]))
    neutral.sort(key=lambda item: (item["name"], item["level"] if item["level"] is not None else -1))
    return providers, consumers, neutral


def numeric_cell(value: Any) -> Any:
    if isinstance(value, float) and math.isclose(value, round(value)):
        return int(round(value))
    return value


def record_row(rank: int, info: Dict[str, Any]) -> List[Any]:
    level = info.get("level")
    return [
        rank,
        info.get("name"),
        info.get("map_type"),
        level if level is not None else "n/a",
        info.get("count"),
        info.get("size_label"),
        numeric_cell(info.get("area")) if info.get("area") is not None else "n/a",
        numeric_cell(info.get("static_pop")),
        numeric_cell(info.get("net_per_building")),
        numeric_cell(info.get("total_net")),
        numeric_cell(info.get("per_tile")) if info.get("area") else "n/a",
    ]


def all_buildings_sort_key(record: Dict[str, Any]) -> Tuple[int, float, str]:
    net_each = float(record["net_per_building"])
    if math.isclose(net_each, 0.0):
        return (2, 0.0, str(record["name"]))
    if net_each > 0:
        return (0, -net_each, str(record["name"]))
    return (1, -net_each, str(record["name"]))


def write_excel_report(
    path: str,
    source_file: str,
    reference_file: str,
    era: str,
    records: List[Dict[str, Any]],
    missing_defs: int,
) -> None:
    providers, consumers, neutral = split_records(records)
    total_provided = sum(record["total_net"] for record in providers)
    total_consumed = -sum(record["total_net"] for record in consumers)
    total_net = total_provided - total_consumed
    placed_count = sum(record["count"] for record in records)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_rows: List[Tuple[str, Any]] = [
        ("Source file", display_path(source_file)),
        ("Reference file", display_path(reference_file) if reference_file != source_file else "same as source"),
        ("Era", era),
        ("Excluded map types", ", ".join(sorted(EXCLUDED_MAP_TYPES))),
        ("Placed building groups", len(records)),
        ("Placed building count", placed_count),
        ("Provider groups", len(providers)),
        ("Consumer groups", len(consumers)),
        ("Neutral groups", len(neutral)),
        ("Population provided", numeric_cell(total_provided)),
        ("Population consumed", numeric_cell(total_consumed)),
        ("Net population", numeric_cell(total_net)),
        ("Placed buildings missing definitions", missing_defs),
        ("Convention", "Positive static population provides population; required population is shown as negative static population."),
    ]
    summary_sheet.append(["Metric", "Value"])
    for row in summary_rows:
        summary_sheet.append(list(row))

    headers = [
        "Rank",
        "Building",
        "Map Type",
        "Level",
        "Count",
        "Size",
        "Area",
        "Static Population",
        "Net Each",
        "Net Total",
        "Net / Tile",
    ]

    sheet_specs = [
        ("Providers", providers, "2E7D68"),
        ("Consumers", consumers, "B55353"),
        ("Neutral", neutral, "6C7A89"),
        ("All Buildings", sorted(records, key=all_buildings_sort_key), "4B6C8B"),
    ]
    for title, sheet_records, tab_color in sheet_specs:
        sheet = workbook.create_sheet(title)
        sheet.sheet_properties.tabColor = tab_color
        sheet.append(headers)
        for idx, info in enumerate(sheet_records, start=1):
            sheet.append(record_row(idx, info))

    apply_workbook_styles(workbook)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    workbook.save(path)


def apply_workbook_styles(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="245C4F")
    summary_fill = PatternFill("solid", fgColor="EAF4F1")
    provider_fill = PatternFill("solid", fgColor="EAF7F1")
    consumer_fill = PatternFill("solid", fgColor="FCEEEF")
    neutral_fill = PatternFill("solid", fgColor="F3F6F8")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True, color="243B35")
    body_font = Font(color="263238")
    positive_font = Font(color="146C43", bold=True)
    negative_font = Font(color="B42318", bold=True)
    thin_side = Side(style="thin", color="D3DED9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row >= 1 and max_col >= 1:
            sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx in range(2, max_row + 1):
            row_fill = alt_fill if row_idx % 2 else PatternFill(fill_type=None)
            if sheet.title == "Summary":
                row_fill = summary_fill
            elif sheet.title == "Providers":
                row_fill = provider_fill if row_idx <= 7 else row_fill
            elif sheet.title == "Consumers":
                row_fill = consumer_fill if row_idx <= 7 else row_fill
            elif sheet.title == "Neutral":
                row_fill = neutral_fill

            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.fill = row_fill
                cell.border = border
                cell.alignment = right if isinstance(cell.value, (int, float)) else left
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not math.isclose(cell.value, round(cell.value)) else '#,##0'

        if sheet.title == "Summary":
            for row_idx in range(2, max_row + 1):
                sheet.cell(row=row_idx, column=1).font = label_font
            sheet.column_dimensions["A"].width = 34
            sheet.column_dimensions["B"].width = 62
            continue

        header_index = {sheet.cell(row=1, column=idx).value: idx for idx in range(1, max_col + 1)}
        for name in ("Static Population", "Net Each", "Net Total", "Net / Tile"):
            col_idx = header_index.get(name)
            if col_idx is None or max_row < 2:
                continue
            col_letter = get_column_letter(col_idx)
            sheet.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{max_row}",
                CellIsRule(operator="greaterThan", formula=["0"], font=positive_font),
            )
            sheet.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{max_row}",
                CellIsRule(operator="lessThan", formula=["0"], font=negative_font),
            )

        widths = {
            "A": 8,
            "B": 34,
            "C": 18,
            "D": 10,
            "E": 10,
            "F": 10,
            "G": 10,
            "H": 18,
            "I": 14,
            "J": 16,
            "K": 13,
        }
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = width


def main() -> None:
    parser = argparse.ArgumentParser(description="Report placed buildings that provide or consume population")
    parser.add_argument(
        "--era",
        default=None,
        help="Fallback era component to inspect. Default: auto-detect from town hall.",
    )
    parser.add_argument(
        "--input",
        default=None,
        help="City JSON file to inspect. Default: latest city_*.json in input/.",
    )
    args = parser.parse_args()

    if not os.path.isdir(INPUT_DIR):
        raise SystemExit(f"Input directory not found: {INPUT_DIR}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    source_file = args.input or latest_city_file()
    city_map, entities, reference_file = load_city_context(source_file)

    chosen_era = args.era or detect_current_era(city_map)
    if not chosen_era:
        raise SystemExit(
            "Could not auto-detect era from CityMapData. Re-run with --era <EraName> (example: --era VirtualFuture)."
        )

    records, missing_defs = collect_population_records(city_map, entities, chosen_era)
    safe_era = chosen_era.replace(" ", "_")
    input_token = safe_output_token(source_file)
    report_path = os.path.join(OUTPUT_DIR, f"building_population_{input_token}_{safe_era}.xlsx")
    write_excel_report(report_path, source_file, reference_file, chosen_era, records, missing_defs)

    providers = sum(1 for record in records if record["total_net"] > 0)
    consumers = sum(1 for record in records if record["total_net"] < 0)
    net_population = sum(record["total_net"] for record in records)

    print(f"Source file: {source_file}")
    if reference_file != source_file:
        print(f"Reference file: {reference_file}")
    print(f"Era inspected: {chosen_era}")
    print(f"Population provider groups: {providers}")
    print(f"Population consumer groups: {consumers}")
    print(f"Net population: {format_number(net_population)}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
