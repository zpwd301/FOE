#!/usr/bin/env python3
"""Generate a comprehensive production report without applying percentage boosts."""
from __future__ import annotations

import argparse
import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from building_attribute_ranking_workbook import (
    PROD_GUILD_GOODS_ATTR,
    PROD_GOODS_ATTR,
    collect_product,
    collect_reward,
    collect_resources,
    component_reward_lookup,
    format_amount,
    is_regular_timed_factory,
    selected_components,
)
from sel_level_fp_report import (
    AGE_BY_LEVEL,
    EntityKey,
    build_entity_index,
    build_rows,
    collect_map_entries,
    default_map_path,
    determine_chain_roles,
    format_time,
    load_json,
    output_basename,
    summarize_available_products,
    summarize_entity_level_production,
    summarize_production_component,
)

FIELDNAMES = [
    "Name",
    "Level",
    "Age",
    "Type",
    "Count",
    "Expected FP Per Collection",
    "Total Expected FP",
    "Expected Medals Per Collection",
    "Total Expected Medals",
    "Expected Goods Per Collection",
    "Total Expected Goods",
    "Expected Guild Goods Per Collection",
    "Total Expected Guild Goods",
    "Expected Troop Units Per Collection",
    "Total Expected Troop Units",
    "Unit Breakdown Per Collection",
    "Passive FP Boost",
    "Entity ID",
    "Production Summary",
]
NUMERIC_FIELDS = {
    "Level",
    "Count",
    "Expected FP Per Collection",
    "Total Expected FP",
    "Expected Medals Per Collection",
    "Total Expected Medals",
    "Expected Goods Per Collection",
    "Total Expected Goods",
    "Expected Guild Goods Per Collection",
    "Total Expected Guild Goods",
    "Expected Troop Units Per Collection",
    "Total Expected Troop Units",
    "Passive FP Boost",
}
UNIT_KEYS = ("prod_unit_current_age", "prod_unit_next_age", "prod_unit_rogue", "prod_units", "prod_resource_unit")
TITLE_FILL = "D6EAF7"
HEADER_FILL = "E7F4DC"
TOP_FILL = "FFF4CC"
ROW_FILLS = ("FFFFFF", "F7FBF4")
BORDER_COLOR = "C7D6E2"
FONT_COLOR = "243447"


@dataclass
class ProductionTotals:
    fp: float = 0.0
    medals: float = 0.0
    goods: float = 0.0
    guild_goods: float = 0.0
    units: float = 0.0
    current_age_units: float = 0.0
    next_age_units: float = 0.0
    rogue_units: float = 0.0

    def add(self, other: "ProductionTotals", factor: float = 1.0) -> None:
        self.fp += other.fp * factor
        self.medals += other.medals * factor
        self.goods += other.goods * factor
        self.guild_goods += other.guild_goods * factor
        self.units += other.units * factor
        self.current_age_units += other.current_age_units * factor
        self.next_age_units += other.next_age_units * factor
        self.rogue_units += other.rogue_units * factor

    def scaled(self, factor: float) -> "ProductionTotals":
        out = ProductionTotals()
        out.add(self, factor)
        return out

    def score(self) -> Tuple[float, float, float, float, float]:
        return (self.fp, self.medals, self.goods, self.guild_goods, self.units)


@dataclass
class ProductionOption:
    label: str
    totals: ProductionTotals


def parse_args() -> argparse.Namespace:
    base_dir = Path(__file__).resolve().parents[1]
    output_dir = base_dir / "output"
    parser = argparse.ArgumentParser(
        description=(
            "Generate a comprehensive city production report for FP, medals, goods, guild goods, and units "
            "without applying percentage production boosts."
        )
    )
    parser.add_argument(
        "--input",
        "--map-file",
        "--sel",
        dest="map_file",
        type=Path,
        default=None,
        help="Path to the map JSON file. Defaults to input/sel or the latest city_*.json.",
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=base_dir / "input" / "ref" / "zpwd-ref",
        help="Reference CityEntities JSON. Defaults to input/ref/zpwd-ref.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=output_dir,
        help="Directory for generated reports.",
    )
    parser.add_argument(
        "--xlsx-only",
        action="store_true",
        help="Write only the Excel workbook and skip keeping the TSV.",
    )
    return parser.parse_args()


def numeric(value: float) -> str:
    if math.isclose(value, round(value)):
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def excel_value(field: str, raw_value: str) -> Any:
    if field not in NUMERIC_FIELDS or raw_value == "":
        return raw_value
    return float(raw_value) if "." in raw_value else int(raw_value)


def option_label(option: Dict[str, Any], fallback: str) -> str:
    parts: List[str] = []
    name = option.get("name")
    if isinstance(name, str) and name:
        parts.append(name)
    for key in ("time", "production_time"):
        time_value = option.get(key)
        if isinstance(time_value, int):
            parts.append(format_time(time_value))
            break
    return " ".join(parts) if parts else fallback


def attrs_to_totals(attrs: Dict[str, float]) -> ProductionTotals:
    current_units = attrs.get("prod_unit_current_age", 0.0)
    next_units = attrs.get("prod_unit_next_age", 0.0)
    rogue_units = attrs.get("prod_unit_rogue", 0.0)
    units = current_units + next_units + rogue_units
    for key in UNIT_KEYS:
        if key not in {"prod_unit_current_age", "prod_unit_next_age", "prod_unit_rogue"}:
            units += attrs.get(key, 0.0)
    return ProductionTotals(
        fp=attrs.get("prod_resource_strategy_points", 0.0),
        medals=attrs.get("prod_resource_medals", 0.0),
        goods=attrs.get(PROD_GOODS_ATTR, 0.0),
        guild_goods=attrs.get(PROD_GUILD_GOODS_ATTR, 0.0),
        units=units,
        current_age_units=current_units,
        next_age_units=next_units,
        rogue_units=rogue_units,
    )


def collect_option_totals(option: Dict[str, Any], reward_lookup: Dict[str, Dict[str, Any]]) -> ProductionTotals:
    attrs: Dict[str, float] = {}
    products = option.get("products")
    if isinstance(products, list):
        for product in products:
            collect_product(attrs, product, 1.0, reward_lookup)
    else:
        product = option.get("product")
        if isinstance(product, dict):
            collect_product(attrs, product, 1.0, reward_lookup)
        reward = option.get("reward")
        if isinstance(reward, dict):
            collect_reward(attrs, reward, 1.0, reward_lookup)
        if not isinstance(product, dict) and not isinstance(reward, dict):
            collect_product(attrs, option, 1.0, reward_lookup)
    return attrs_to_totals(attrs)


def collect_level_product_totals(production_value: Dict[str, Any]) -> ProductionTotals:
    attrs: Dict[str, float] = {}
    resource_type = production_value.get("type")
    value = production_value.get("value")
    if isinstance(resource_type, str) and isinstance(value, (int, float)):
        collect_resources(attrs, "prod_resource", {"resources": {resource_type: value}})
    return attrs_to_totals(attrs)


def production_options(entity_def: Dict[str, Any], era_name: str) -> List[ProductionOption]:
    options: List[ProductionOption] = []
    for component_name, component in selected_components(entity_def, era_name):
        production = component.get("production")
        if not isinstance(production, dict):
            continue
        raw_options = production.get("options")
        if not isinstance(raw_options, list):
            continue
        reward_lookup = component_reward_lookup(component)
        for idx, option in enumerate(raw_options, start=1):
            if not isinstance(option, dict):
                continue
            options.append(
                ProductionOption(
                    label=option_label(option, f"{component_name} option {idx}"),
                    totals=collect_option_totals(option, reward_lookup),
                )
            )

    available_products = entity_def.get("available_products")
    if isinstance(available_products, list) and not is_regular_timed_factory(entity_def):
        for idx, option in enumerate(available_products, start=1):
            if not isinstance(option, dict):
                continue
            options.append(
                ProductionOption(
                    label=option_label(option, f"available product {idx}"),
                    totals=collect_option_totals(option, {}),
                )
            )

    levels = entity_def.get("entity_levels")
    available_products = entity_def.get("available_products")
    if isinstance(levels, list) and isinstance(available_products, list):
        era_entry = next((entry for entry in levels if isinstance(entry, dict) and entry.get("era") == era_name), None)
        production_values = era_entry.get("production_values") if isinstance(era_entry, dict) else None
        if isinstance(production_values, list):
            for idx, production_value in enumerate(production_values):
                if not isinstance(production_value, dict):
                    continue
                label = f"level product {idx + 1}"
                if idx < len(available_products) and isinstance(available_products[idx], dict):
                    label = option_label(available_products[idx], label)
                options.append(ProductionOption(label=label, totals=collect_level_product_totals(production_value)))

    return options


def best_totals(options: Sequence[ProductionOption]) -> ProductionTotals:
    best = ProductionTotals()
    for option in options:
        best.fp = max(best.fp, option.totals.fp)
        best.medals = max(best.medals, option.totals.medals)
        best.goods = max(best.goods, option.totals.goods)
        best.guild_goods = max(best.guild_goods, option.totals.guild_goods)
        best.units = max(best.units, option.totals.units)
        best.current_age_units = max(best.current_age_units, option.totals.current_age_units)
        best.next_age_units = max(best.next_age_units, option.totals.next_age_units)
        best.rogue_units = max(best.rogue_units, option.totals.rogue_units)
    return best


def unit_breakdown(totals: ProductionTotals) -> str:
    parts = []
    if totals.current_age_units:
        parts.append(f"current age={numeric(totals.current_age_units)}")
    if totals.next_age_units:
        parts.append(f"next age={numeric(totals.next_age_units)}")
    if totals.rogue_units:
        parts.append(f"rogue={numeric(totals.rogue_units)}")
    if totals.units and not parts:
        parts.append(f"units={numeric(totals.units)}")
    return ", ".join(parts)


def option_summary(options: Sequence[ProductionOption]) -> str:
    parts: List[str] = []
    for option in options:
        values = []
        if option.totals.fp:
            values.append(f"{format_amount(option.totals.fp)} FP")
        if option.totals.medals:
            values.append(f"{format_amount(option.totals.medals)} medals")
        if option.totals.goods:
            values.append(f"{format_amount(option.totals.goods)} goods")
        if option.totals.guild_goods:
            values.append(f"{format_amount(option.totals.guild_goods)} guild goods")
        if option.totals.units:
            values.append(f"{format_amount(option.totals.units)} units")
        if values:
            parts.append(f"{option.label}: {', '.join(values)}")
    return " | ".join(parts)


def chain_piece_totals(entity_def: Dict[str, Any]) -> ProductionTotals:
    attrs: Dict[str, float] = {}
    chain = entity_def.get("components", {}).get("AllAge", {}).get("chain", {})
    bonuses = chain.get("config", {}).get("bonuses")
    if not isinstance(bonuses, list):
        return ProductionTotals()
    for bonus in bonuses:
        productions = bonus.get("productions") if isinstance(bonus, dict) else None
        if not isinstance(productions, list):
            continue
        for production in productions:
            collect_product(attrs, production, 1.0, {})
    return attrs_to_totals(attrs)


def chain_bonus_totals(
    counts: Dict[EntityKey, int],
    entity_defs: Dict[str, Dict[str, Any]],
) -> DefaultDict[EntityKey, ProductionTotals]:
    bonuses: DefaultDict[EntityKey, ProductionTotals] = defaultdict(ProductionTotals)
    chain_mains, chain_pieces = determine_chain_roles(list(counts), entity_defs)
    for chain_id, pieces in chain_pieces.items():
        mains = chain_mains.get(chain_id, [])
        if len(mains) != 1:
            continue
        main_key = mains[0]
        for piece_key in pieces:
            piece_def = entity_defs.get(piece_key.cityentity_id, {})
            bonuses[main_key].add(chain_piece_totals(piece_def), counts[piece_key])
    return bonuses


def production_summary(
    entity_def: Dict[str, Any],
    era_name: str,
    name_by_id: Dict[str, str],
    options: Sequence[ProductionOption],
    fallback_summary: str,
) -> str:
    summary_parts: List[str] = []
    numeric_summary = option_summary(options)
    if numeric_summary:
        summary_parts.append(numeric_summary)
    reference_summary = summarize_production_component(entity_def, era_name, name_by_id)
    if reference_summary is None:
        reference_summary = summarize_entity_level_production(entity_def, era_name)
    if reference_summary is None:
        reference_summary = summarize_available_products(entity_def)
    if reference_summary and reference_summary not in summary_parts:
        summary_parts.append(reference_summary)
    if not summary_parts:
        summary_parts.append(fallback_summary)
    return " || ".join(summary_parts)


def build_report_rows(
    counts: Dict[EntityKey, int],
    entity_defs: Dict[str, Dict[str, Any]],
    name_by_id: Dict[str, str],
) -> List[Dict[str, str]]:
    fp_rows = {row.key: row for row in build_rows(counts, entity_defs, name_by_id)}
    chain_bonuses = chain_bonus_totals(counts, entity_defs)
    rows: List[Dict[str, str]] = []
    for key in sorted(counts, key=lambda item: (name_by_id.get(item.cityentity_id, item.cityentity_id), item.level, item.entity_type)):
        count = counts[key]
        age_code, era_name = AGE_BY_LEVEL.get(key.level, (f"L{key.level}", ""))
        entity_def = entity_defs.get(key.cityentity_id, {})
        name = str(entity_def.get("name", key.cityentity_id))
        options = production_options(entity_def, era_name)
        per_collection = best_totals(options)
        total = per_collection.scaled(count)
        total.add(chain_bonuses[key])
        if count:
            per_collection = total.scaled(1 / count)
        fp_row = fp_rows.get(key)
        fallback_summary = fp_row.production if fp_row else "no production found in input/ref/zpwd-ref for this age"
        summary = production_summary(entity_def, era_name, name_by_id, options, fallback_summary)
        if chain_bonuses[key].score() != (0.0, 0.0, 0.0, 0.0, 0.0):
            summary += (
                " || chain bonus counted on this row: "
                f"{numeric(chain_bonuses[key].fp)} FP, "
                f"{numeric(chain_bonuses[key].medals)} medals, "
                f"{numeric(chain_bonuses[key].goods)} goods, "
                f"{numeric(chain_bonuses[key].guild_goods)} guild goods, "
                f"{numeric(chain_bonuses[key].units)} units"
            )
        rows.append(
            {
                "Name": name,
                "Level": str(key.level),
                "Age": age_code,
                "Type": key.entity_type,
                "Count": str(count),
                "Expected FP Per Collection": numeric(per_collection.fp),
                "Total Expected FP": numeric(total.fp),
                "Expected Medals Per Collection": numeric(per_collection.medals),
                "Total Expected Medals": numeric(total.medals),
                "Expected Goods Per Collection": numeric(per_collection.goods),
                "Total Expected Goods": numeric(total.goods),
                "Expected Guild Goods Per Collection": numeric(per_collection.guild_goods),
                "Total Expected Guild Goods": numeric(total.guild_goods),
                "Expected Troop Units Per Collection": numeric(per_collection.units),
                "Total Expected Troop Units": numeric(total.units),
                "Unit Breakdown Per Collection": unit_breakdown(per_collection),
                "Passive FP Boost": numeric(fp_row.passive_fp_boost if fp_row else 0.0),
                "Entity ID": key.cityentity_id,
                "Production Summary": summary,
            }
        )
    rows.sort(
        key=lambda row: (
            -float(row["Total Expected FP"] or 0),
            -float(row["Total Expected Medals"] or 0),
            -float(row["Total Expected Goods"] or 0),
            -float(row["Total Expected Guild Goods"] or 0),
            -float(row["Total Expected Troop Units"] or 0),
            row["Name"],
        )
    )
    return rows


def write_tsv(rows: Sequence[Dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def city_name_from_input(map_file: Path) -> str:
    name = output_basename(map_file).replace("_", " ").replace("-", " ").strip()
    return name.title() if name else "City"


def write_excel(rows: Sequence[Dict[str, str]], output_path: Path, map_file: Path) -> None:
    city_name = city_name_from_input(map_file)
    source_name = map_file.name
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    details = workbook.create_sheet("Building Production")
    workbook.create_sheet("Notes")

    totals = {field: sum(float(row[field] or 0) for row in rows) for field in FIELDNAMES if field.startswith("Total")}
    summary_values = [
        ("Source", source_name),
        ("Placed building rows", len(rows)),
        ("Total expected FP", totals["Total Expected FP"]),
        ("Total expected medals", totals["Total Expected Medals"]),
        ("Total expected goods", totals["Total Expected Goods"]),
        ("Total expected guild goods", totals["Total Expected Guild Goods"]),
        ("Total expected troop units", totals["Total Expected Troop Units"]),
    ]
    summary["A1"] = f"{city_name} City Comprehensive Production Summary"
    summary["A1"].font = Font(bold=True, size=16, color=FONT_COLOR)
    summary["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    summary.merge_cells("A1:B1")
    for row_idx, (label, value) in enumerate(summary_values, start=3):
        summary.cell(row_idx, 1, label)
        summary.cell(row_idx, 2, value)
        summary.cell(row_idx, 1).font = Font(bold=True, color=FONT_COLOR)
        if isinstance(value, (int, float)):
            summary.cell(row_idx, 2).number_format = "#,##0.00"
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 24
    summary.sheet_view.showGridLines = False
    summary.sheet_properties.tabColor = "5B9BD5"

    for col_idx, header in enumerate(FIELDNAMES, start=1):
        details.cell(1, col_idx, header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(FIELDNAMES, start=1):
            details.cell(row_idx, col_idx, excel_value(header, row[header]))
    apply_detail_styles(details, len(rows))

    notes = workbook["Notes"]
    notes["A1"] = "Report Notes"
    notes["A1"].font = Font(bold=True, size=15, color=FONT_COLOR)
    notes["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    notes["A3"] = f"Input file: {source_name}"
    notes["A4"] = f"Report generated: {generated_at}"
    notes["A6"] = "Expected values use reward drop chances when available."
    notes["A7"] = "When a building has multiple production options, each category column uses that building's best option for that category."
    notes["A8"] = "Chain-piece bonuses are counted on the matching main chain building when a single main building is present."
    notes.column_dimensions["A"].width = 120
    notes.sheet_view.showGridLines = False
    notes.sheet_properties.tabColor = "A6A6A6"

    workbook.save(output_path)


def apply_detail_styles(sheet, row_count: int) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{row_count + 1}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "70AD47"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=FONT_COLOR)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_idx in range(2, row_count + 2):
        fill = PatternFill("solid", fgColor=ROW_FILLS[row_idx % 2])
        for col_idx, header in enumerate(FIELDNAMES, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if header in NUMERIC_FIELDS else "left",
                vertical="top",
                wrap_text=header in {"Production Summary", "Unit Breakdown Per Collection"},
            )
            if header in NUMERIC_FIELDS:
                cell.number_format = "#,##0.00"
        if row_idx <= 6:
            for col_idx in range(1, len(FIELDNAMES) + 1):
                sheet.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=TOP_FILL)
    widths = {
        "Name": 34,
        "Level": 8,
        "Age": 9,
        "Type": 18,
        "Count": 8,
        "Unit Breakdown Per Collection": 28,
        "Entity ID": 38,
        "Production Summary": 110,
    }
    for col_idx, header in enumerate(FIELDNAMES, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 18)
    if row_count:
        for header in (
            "Total Expected FP",
            "Total Expected Medals",
            "Total Expected Goods",
            "Total Expected Guild Goods",
            "Total Expected Troop Units",
        ):
            col_idx = FIELDNAMES.index(header) + 1
            sheet.conditional_formatting.add(
                f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{row_count + 1}",
                ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="B7D7F0"),
            )


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parents[1]
    input_dir = base_dir / "input"
    map_file = args.map_file if args.map_file is not None else default_map_path(input_dir)
    map_data = load_json(map_file)
    reference_data = load_json(args.reference)
    entity_defs, name_by_id = build_entity_index(reference_data)
    counts = collect_map_entries(map_data)
    rows = build_report_rows(counts, entity_defs, name_by_id)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    base_name = output_basename(map_file)
    tsv_path = args.output_dir / f"{base_name}_comprehensive_production_without_percentage_boost_report.tsv"
    xlsx_path = args.output_dir / f"{base_name}_comprehensive_production_without_percentage_boost_report.xlsx"
    if not args.xlsx_only:
        write_tsv(rows, tsv_path)
    write_excel(rows, xlsx_path, map_file)
    print(f"Wrote {xlsx_path}")
    if not args.xlsx_only:
        print(f"Wrote {tsv_path}")


if __name__ == "__main__":
    main()
