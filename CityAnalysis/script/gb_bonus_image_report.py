#!/usr/bin/env python3
"""Extract GB bonus tables from images and export an Excel report."""
from __future__ import annotations

import argparse
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
from rapidocr_onnxruntime import RapidOCR

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = BASE_DIR / "input" / "GB-bonus-images"
DEFAULT_OUTPUT_PATH = BASE_DIR / "output" / "gb_bonus_from_images.xlsx"
TARGET_LEVELS = [1, 101, 201, 301, 400]


@dataclass
class OcrToken:
    text: str
    score: float
    cx: float
    cy: float


@dataclass
class RowResult:
    boost_kind: str
    lv1: str
    lv101: str
    lv201: str
    lv301: str
    lv400: str
    icon_png: Optional[bytes] = None


@dataclass
class ImageResult:
    image_file: str
    gb_name: str
    rows: List[RowResult]


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def to_tokens(ocr_result: Sequence[Sequence[object]]) -> List[OcrToken]:
    out: List[OcrToken] = []
    for item in ocr_result:
        if len(item) != 3:
            continue
        points, text, score = item
        if not isinstance(points, list) or not isinstance(text, str):
            continue
        if len(points) != 4:
            continue

        xs = [float(p[0]) for p in points]
        ys = [float(p[1]) for p in points]
        cx = sum(xs) / 4.0
        cy = sum(ys) / 4.0
        try:
            numeric_score = float(score)
        except Exception:  # noqa: BLE001
            numeric_score = 0.0

        out.append(OcrToken(text=normalize_space(text), score=numeric_score, cx=cx, cy=cy))
    return out


def split_camel(text: str) -> str:
    text = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Za-z])(?=[0-9])", " ", text)
    text = re.sub(r"(?<=[0-9])(?=[A-Za-z])", " ", text)
    return normalize_space(text)


def extract_gb_name(tokens: Sequence[OcrToken], image_stem: str) -> str:
    top_tokens = [t for t in tokens if t.cy < 80]
    bonus_title = None

    for token in sorted(top_tokens, key=lambda t: (-t.score, t.cy)):
        compact = re.sub(r"[^A-Za-z]", "", token.text).lower()
        if "bonusoverview" in compact:
            bonus_title = token.text
            break

    if bonus_title is None:
        return split_camel(image_stem)

    raw = bonus_title
    raw = re.sub(r"(?i)bonus\s*overview", "", raw)
    raw = re.sub(r"(?i)bonusoverview", "", raw)
    raw = split_camel(raw)

    if not raw:
        return split_camel(image_stem)
    return raw


def parse_level_from_token(text: str) -> Optional[int]:
    compact = re.sub(r"\s+", "", text).lower()
    compact = compact.replace("l", "l")
    if "level" not in compact:
        return None

    digits = re.sub(r"\D", "", compact)
    if not digits:
        return None

    value = int(digits)
    if value in TARGET_LEVELS:
        return value
    return None


def detect_level_columns(tokens: Sequence[OcrToken], width: int, height: int) -> Tuple[List[float], float]:
    level_hits: Dict[int, OcrToken] = {}
    for token in tokens:
        level = parse_level_from_token(token.text)
        if level is None:
            continue
        prev = level_hits.get(level)
        if prev is None or token.score > prev.score:
            level_hits[level] = token

    if len(level_hits) >= 2:
        points = []
        for idx, level in enumerate(TARGET_LEVELS):
            token = level_hits.get(level)
            if token is not None:
                points.append((idx, token.cx))

        n = len(points)
        sum_i = sum(float(idx) for idx, _ in points)
        sum_x = sum(x for _, x in points)
        sum_ii = sum(float(idx * idx) for idx, _ in points)
        sum_ix = sum(float(idx) * x for idx, x in points)

        denom = n * sum_ii - sum_i * sum_i
        if abs(denom) > 1e-6:
            slope = (n * sum_ix - sum_i * sum_x) / denom
            intercept = (sum_x - slope * sum_i) / n
            centers = [intercept + slope * i for i in range(5)]
            header_y = max(level_hits[level].cy for level in level_hits)
            return centers, header_y

    fallback = [0.22 * width, 0.39 * width, 0.56 * width, 0.73 * width, 0.90 * width]
    header_y = 0.42 * height
    return fallback, header_y


def normalize_cell_text(text: str) -> str:
    value = text.strip()
    if not value:
        return "-"

    lowered = value.lower().replace(" ", "")
    if lowered in {"-", "_", "--"}:
        return "-"

    # OCR often reads '/24h' as '124h' or '|24h'.
    value = re.sub(r"^[|1]?24h$", "/24h", value, flags=re.IGNORECASE)
    return value


def cluster_rows(y_values: Sequence[float], threshold: float) -> List[float]:
    if not y_values:
        return []

    sorted_y = sorted(y_values)
    clusters: List[List[float]] = [[sorted_y[0]]]
    for y in sorted_y[1:]:
        current = clusters[-1]
        current_center = sum(current) / len(current)
        if y - current_center <= threshold:
            current.append(y)
        else:
            clusters.append([y])

    return [sum(cluster) / len(cluster) for cluster in clusters]


def row_boundaries(row_centers: Sequence[float], header_y: float, height: int) -> List[Tuple[int, int]]:
    if not row_centers:
        return []
    bounds: List[Tuple[int, int]] = []
    for idx, center in enumerate(row_centers):
        if idx == 0:
            top = int(max(header_y + 10, center - 30))
        else:
            top = int((row_centers[idx - 1] + center) / 2.0)

        if idx == len(row_centers) - 1:
            bottom = int(min(height - 2, center + 30))
        else:
            bottom = int((center + row_centers[idx + 1]) / 2.0)

        if bottom <= top:
            bottom = top + 1
        bounds.append((top, bottom))
    return bounds


def extract_rows(tokens: Sequence[OcrToken], image: Image.Image) -> List[RowResult]:
    width, height = image.size
    col_centers, header_y = detect_level_columns(tokens, width, height)

    gap = (col_centers[-1] - col_centers[0]) / 4.0
    x_min = col_centers[0] - 0.55 * gap
    x_max = col_centers[-1] + 0.55 * gap

    value_tokens = []
    for token in tokens:
        if token.cy <= header_y + 12:
            continue
        if token.cx < x_min or token.cx > x_max:
            continue
        if parse_level_from_token(token.text) is not None:
            continue
        value_tokens.append(token)

    if not value_tokens:
        return []

    row_threshold = max(16.0, 0.055 * height)
    row_centers = cluster_rows([token.cy for token in value_tokens], threshold=row_threshold)
    if not row_centers:
        return []
    bounds = row_boundaries(row_centers, header_y, height)

    token_row_index: Dict[int, int] = {}
    for idx, token in enumerate(value_tokens):
        nearest = min(range(len(row_centers)), key=lambda i: abs(token.cy - row_centers[i]))
        if abs(token.cy - row_centers[nearest]) <= row_threshold:
            token_row_index[idx] = nearest

    col_threshold = 0.60 * gap
    rows: List[RowResult] = []
    for row_idx, row_center in enumerate(row_centers, start=1):
        row_data = {
            "lv1": "-",
            "lv101": "-",
            "lv201": "-",
            "lv301": "-",
            "lv400": "-",
        }

        for level_idx, level in enumerate(TARGET_LEVELS):
            center_x = col_centers[level_idx]
            cell_tokens = [
                token
                for token_index, token in enumerate(value_tokens)
                if token_row_index.get(token_index) == (row_idx - 1) and abs(token.cx - center_x) <= col_threshold
            ]
            if not cell_tokens:
                continue

            cell_tokens.sort(key=lambda item: item.cy)
            merged = "\n".join(normalize_cell_text(token.text) for token in cell_tokens)
            row_data[f"lv{level}"] = normalize_cell_text(merged)

        # Keep rows that have at least one non-empty value.
        if not any(row_data[f"lv{level}"] != "-" for level in TARGET_LEVELS):
            continue

        icon_x1 = int(max(0, col_centers[0] - 1.45 * gap))
        icon_x2 = int(max(icon_x1 + 1, min(width, col_centers[0] - 0.55 * gap)))
        top, bottom = bounds[row_idx - 1]
        crop = image.crop((icon_x1, max(0, top + 1), icon_x2, min(height, bottom - 1)))
        icon_png: Optional[bytes] = None
        if crop.width > 1 and crop.height > 1:
            icon_buffer = io.BytesIO()
            crop.save(icon_buffer, format="PNG")
            icon_png = icon_buffer.getvalue()

        rows.append(
            RowResult(
                boost_kind=f"Boost {row_idx}",
                lv1=row_data["lv1"],
                lv101=row_data["lv101"],
                lv201=row_data["lv201"],
                lv301=row_data["lv301"],
                lv400=row_data["lv400"],
                icon_png=icon_png,
            )
        )

    return rows


def process_image(path: Path, ocr: RapidOCR) -> ImageResult:
    ocr_result, _ = ocr(str(path))
    if not ocr_result:
        return ImageResult(image_file=path.name, gb_name=split_camel(path.stem), rows=[])

    tokens = to_tokens(ocr_result)

    with Image.open(path) as image:
        rgb_image = image.convert("RGB")

    gb_name = extract_gb_name(tokens, path.stem)
    rows = extract_rows(tokens, image=rgb_image)
    return ImageResult(image_file=path.name, gb_name=gb_name, rows=rows)


def write_excel(results: Sequence[ImageResult], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GB Bonus"

    headers = [
        "gb_name",
        "boost_kind",
        "lv1",
        "lv101",
        "lv201",
        "lv301",
        "lv400",
    ]
    sheet.append(headers)

    body_font = Font(name="Calibri", size=11)
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="5A2D0C", end_color="5A2D0C")
    thin = Side(style="thin", color="C8B08B")
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    name_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = body_border

    row_icon_images: List[XLImage] = []
    row_icon_streams: List[io.BytesIO] = []
    group_ranges: List[Tuple[int, int]] = []
    for result in sorted(results, key=lambda item: item.gb_name.lower()):
        group_start = sheet.max_row + 1
        if not result.rows:
            sheet.append([result.gb_name, "(no rows parsed)", "-", "-", "-", "-", "-"])
            group_ranges.append((group_start, group_start))
            continue

        for row in result.rows:
            excel_row = sheet.max_row + 1
            sheet.append(
                [
                    result.gb_name,
                    "",
                    row.lv1,
                    row.lv101,
                    row.lv201,
                    row.lv301,
                    row.lv400,
                ]
            )
            if row.icon_png is not None:
                icon_stream = io.BytesIO(row.icon_png)
                icon = XLImage(icon_stream)
                target_h = 30.0
                if icon.height > 0:
                    ratio = target_h / float(icon.height)
                    icon.width = max(1, int(icon.width * ratio))
                    icon.height = max(1, int(icon.height * ratio))
                icon.anchor = f"B{excel_row}"
                sheet.add_image(icon)
                row_icon_images.append(icon)
                row_icon_streams.append(icon_stream)
                sheet.row_dimensions[excel_row].height = max(sheet.row_dimensions[excel_row].height or 15, 28)
            else:
                sheet.cell(excel_row, 2).value = row.boost_kind
        group_ranges.append((group_start, sheet.max_row))

    for row in sheet.iter_rows(min_row=2, max_col=7):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            if cell.column == 1:
                cell.alignment = name_alignment
            else:
                cell.alignment = center

    for start, end in group_ranges:
        if end > start:
            sheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            merged_cell = sheet.cell(start, 1)
            merged_cell.alignment = name_alignment
            merged_cell.font = body_font
            merged_cell.border = body_border

    widths = {
        "A": 28,
        "B": 12,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="OCR GB bonus overview images and export level values to Excel"
    )
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_DIR),
        help=f"Input image directory. Default: {DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help=f"Output Excel path. Default: {DEFAULT_OUTPUT_PATH}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not input_dir.is_dir():
        raise SystemExit(f"Input directory not found: {input_dir}")

    image_paths = sorted([p for p in input_dir.iterdir() if p.suffix.lower() in {'.png', '.jpg', '.jpeg', '.webp'}])
    if not image_paths:
        raise SystemExit(f"No image files found in: {input_dir}")

    ocr = RapidOCR()
    results = [process_image(path, ocr) for path in image_paths]
    write_excel(results, output_path)

    parsed_rows = sum(len(result.rows) for result in results)
    print(f"Images processed: {len(results)}")
    print(f"Rows extracted: {parsed_rows}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
